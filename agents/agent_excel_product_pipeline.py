#!/usr/bin/env python3
"""Excel-driven newest-product Draft pipeline for RADMAN SILVER.

Excel is the sole catalog-data source. Public legacy pages are consulted only to
locate and download original gallery images. Imports are create-only Drafts.
"""

from __future__ import annotations

import argparse
import base64
import csv
import hashlib
import html
import io
import json
import mimetypes
import os
import re
import sys
import urllib.error
import urllib.parse
from collections import Counter
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, ROUND_FLOOR
from html.parser import HTMLParser
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

REPO_ROOT = Path(__file__).resolve().parent.parent
sys.path.insert(0, str(REPO_ROOT))

from agents.agent_legacy_catalog_pilot import (  # noqa: E402
    BASE_URL,
    MAX_IMAGE_BYTES,
    RateLimitedFetcher,
    iri_to_uri,
    parse_product_links,
    parse_sitemap_urls,
)
from agents.agent_original_image_processor import process_product  # noqa: E402
from agents.lib.legacy_pricing import round_up_toman  # noqa: E402
from agents.lib.product_identity import (  # noqa: E402
    build_legacy_identity_key,
    clean_public_product_title,
    normalize_identity_digits,
)
from scripts.analyze_excel_catalog import (  # noqa: E402
    COL_ACTIVE,
    COL_AVAILABILITY,
    COL_CATEGORY,
    COL_ID,
    COL_PRICE,
    COL_TITLE,
    COL_WEIGHT,
    SHEET_NAME,
    display_text,
    is_blank,
    normalize_availability,
    normalize_text,
    parse_active,
    parse_decimal,
    parse_id,
)
from scripts.import_products import WPGateway, WPCliError  # noqa: E402

DEFAULT_EXCEL_FILE = Path(
    "/home/radmansi/radman-deploy/products_20260821_182238.xlsx"
)
DEFAULT_MAX_PRODUCTS = 1000
HARD_MAX_PRODUCTS = 1000
EXPECTED_APP_ENV = "staging"
EXPECTED_WP_URL = "https://staging.radmansilver.ir"
EXPECTED_WP_PATH = "/home/radmansi/staging.radmansilver.ir"
REQUIRED_CURRENCY = "IRT"
API_SLOT = Path("/home/radmansi/.config/radman/api-keys/legacy-site.env")
PIPELINE_VERSION = "PR-30A"
EXCEL_IMAGE_USER_AGENT = (
    "RadmanSilverExcelImageImporter/1.0 "
    "(+https://radmansilver.ir; owner-controlled original-gallery migration)"
)
TEHRAN = ZoneInfo("Asia/Tehran")

COL_PRE_DISCOUNT_PRICE = 10
COL_STOCK = 11
COL_RAW_CODE = 27
COL_SEO_TITLE = 28
COL_SEO_DESCRIPTION = 29
MIN_REQUIRED_COLUMNS = 29

STANDARD_RATE = 650_000
LARGE_STONE_RATE = 590_000
ROUNDING_STEP = 50_000
LARGE_STONE_RE = re.compile(
    r"(?:(?:نگین|عقیق).{0,20}(?:درشت|بزرگ)|(?:درشت|بزرگ).{0,20}(?:نگین|عقیق))"
)
NO_STONE_RE = re.compile(r"(?:بدون|فاقد|بی)\s*(?:نگین|سنگ)")

REPORT_COLUMNS = (
    "wp_id",
    "old_public_title",
    "new_public_title",
    "title_cleanup_applied",
    "extracted_title_code",
    "current_sku",
    "sku_title_match",
    "legacy_product_id",
    "legacy_raw_code",
    "legacy_url",
    "identity_key",
    "title_cleanup_status",
    "description_updated",
    "specs_found_count",
    "price_changed",
    "legacy_id",
    "sku",
    "sku_source",
    "title",
    "category_raw",
    "category",
    "weight_grams",
    "excel_price_toman",
    "pre_discount_price_toman",
    "computed_price_toman",
    "final_price_toman",
    "regular_price_toman",
    "price_source",
    "rate_used",
    "stone_class",
    "stone_type",
    "stone_color",
    "band_type",
    "silver_purity",
    "spec_weight",
    "weight_source",
    "description_source",
    "unknown_spec_labels_seen",
    "stock",
    "images_found",
    "image_status",
    "image_discovery_strategy",
    "action",
    "review_flags",
    "wordpress_product_id",
)


class ExcelPipelineError(RuntimeError):
    pass


@dataclass(frozen=True)
class PricingDecision:
    stone_class: str
    rate_used: int
    excel_price_toman: Optional[int]
    computed_price_toman: Optional[Decimal]
    final_price_toman: Optional[int]
    pre_discount_price_toman: Optional[int]
    regular_price_toman: Optional[int]
    price_source: str
    review_flags: Tuple[str, ...]

    def to_dict(self) -> Dict[str, Any]:
        value = asdict(self)
        if self.computed_price_toman is not None:
            value["computed_price_toman"] = format(
                self.computed_price_toman, "f"
            )
        value["review_flags"] = list(self.review_flags)
        return value


@dataclass(frozen=True)
class SKUDecision:
    sku: str
    source: str
    raw_code: str


KNOWN_SPEC_LABELS = {
    "دسته بندی": "category",
    "وزن": "weight",
    "نوع رکاب": "band_type",
    "رنگ نگین": "stone_color",
    "نوع سنگ": "stone_type",
    "نوع حکاکی": "engraving_type",
    "عیار نقره": "silver_purity",
    "سایز": "size",
}


@dataclass(frozen=True)
class LegacySpecs:
    all_specs: Dict[str, str]
    category: str
    weight_display: str
    weight_grams: Optional[Decimal]
    band_type: str
    stone_color: str
    stone_type: str
    engraving_type: str
    silver_purity: str
    size: str
    unknown_labels: Tuple[str, ...]

    @property
    def found(self) -> bool:
        return bool(self.all_specs)

    def to_dict(self) -> Dict[str, Any]:
        value = asdict(self)
        value["weight_grams"] = (
            format(self.weight_grams, "f") if self.weight_grams is not None else None
        )
        value["unknown_labels"] = list(self.unknown_labels)
        return value


@dataclass(frozen=True)
class LegacyPageResult:
    url: str
    image_urls: Tuple[str, ...]
    strategy: str
    specs: LegacySpecs


class SpecTextHTMLParser(HTMLParser):
    """Render HTML to text and preserve dl/table label-value pairs."""

    BLOCK_TAGS = {
        "article",
        "br",
        "dd",
        "div",
        "dl",
        "dt",
        "h1",
        "h2",
        "h3",
        "h4",
        "li",
        "p",
        "section",
        "table",
        "td",
        "th",
        "tr",
        "ul",
    }

    def __init__(self) -> None:
        super().__init__(convert_charrefs=True)
        self.parts: List[str] = []
        self.structured_pairs: List[Tuple[str, str]] = []
        self._ignored_depth = 0
        self._capture_tag: Optional[str] = None
        self._capture_parts: List[str] = []
        self._pending_label: Optional[str] = None

    def handle_starttag(
        self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]
    ) -> None:
        if tag in {"script", "style", "noscript"}:
            self._ignored_depth += 1
            return
        if self._ignored_depth:
            return
        if tag in self.BLOCK_TAGS:
            self.parts.append("\n")
        if tag in {"dt", "dd", "th", "td"}:
            self._capture_tag = tag
            self._capture_parts = []

    def handle_endtag(self, tag: str) -> None:
        if tag in {"script", "style", "noscript"} and self._ignored_depth:
            self._ignored_depth -= 1
            return
        if self._ignored_depth:
            return
        if self._capture_tag == tag:
            text = normalize_text(" ".join(self._capture_parts))
            if tag in {"dt", "th"}:
                self._pending_label = text or None
            elif tag == "td" and text:
                if self._pending_label:
                    self.structured_pairs.append((self._pending_label, text))
                    self._pending_label = None
                else:
                    self._pending_label = text.strip(" :：") or None
            elif tag == "dd" and self._pending_label and text:
                self.structured_pairs.append((self._pending_label, text))
                self._pending_label = None
            self._capture_tag = None
            self._capture_parts = []
        if tag in self.BLOCK_TAGS:
            self.parts.append("\n")

    def handle_data(self, data: str) -> None:
        if self._ignored_depth:
            return
        text = html.unescape(data)
        self.parts.append(text)
        if self._capture_tag:
            self._capture_parts.append(text)

    @property
    def text(self) -> str:
        return "".join(self.parts)


class GalleryHTMLParser(HTMLParser):
    """Collect ordered same-page image candidates without parsing product data."""

    def __init__(self, page_url: str) -> None:
        super().__init__(convert_charrefs=True)
        self.page_url = page_url
        self.candidates: List[str] = []

    @staticmethod
    def _attrs(attrs: Sequence[Tuple[str, Optional[str]]]) -> Dict[str, str]:
        return {key.lower(): value or "" for key, value in attrs}

    def _add(self, value: str) -> None:
        value = html.unescape(str(value or "")).strip()
        if value:
            self.candidates.append(urllib.parse.urljoin(self.page_url, value))

    def handle_starttag(
        self, tag: str, attrs: Sequence[Tuple[str, Optional[str]]]
    ) -> None:
        values = self._attrs(attrs)
        if tag == "img":
            for key in (
                "data-large_image",
                "data-zoom-image",
                "data-src",
                "data-lazy-src",
                "src",
            ):
                if values.get(key):
                    self._add(values[key])
                    break
            srcset = values.get("srcset") or values.get("data-srcset")
            if srcset:
                entries = []
                for item in srcset.split(","):
                    parts = item.strip().split()
                    if not parts:
                        continue
                    width = 0
                    if len(parts) > 1 and parts[1].lower().endswith("w"):
                        try:
                            width = int(parts[1][:-1])
                        except ValueError:
                            width = 0
                    entries.append((width, parts[0]))
                if entries:
                    self._add(max(entries, key=lambda item: item[0])[1])
        elif tag == "a":
            href = values.get("href", "")
            if re.search(r"\.(?:jpe?g|png|webp)(?:[?#]|$)", href, re.I):
                self._add(href)
        elif tag == "meta":
            key = (values.get("property") or values.get("name") or "").lower()
            if key in {"og:image", "twitter:image"}:
                self._add(values.get("content", ""))


def _cell(values: Sequence[Any], one_based_column: int) -> Any:
    index = one_based_column - 1
    return values[index] if index < len(values) else None


def _raw_trace(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _positive_toman(value: Any) -> Optional[int]:
    parsed = parse_decimal(value)
    if parsed is None or parsed <= 0:
        return None
    return int(parsed.to_integral_value(rounding=ROUND_FLOOR))


def parse_weight(value: Any) -> Optional[Decimal]:
    parsed = parse_decimal(value, decimal_comma=True)
    if parsed is None or parsed <= 0 or parsed > Decimal("500"):
        return None
    return parsed


def empty_legacy_specs() -> LegacySpecs:
    return LegacySpecs({}, "", "", None, "", "", "", "", "", "", tuple())


def _normalized_spec_label(label: str) -> str:
    value = normalize_text(label).strip(" .،؛:：-")
    if value.startswith("مشخصات"):
        value = normalize_text(value[len("مشخصات") :]).strip(" .،؛:：-")
    return value


def legacy_specs_from_pairs(pairs: Iterable[Tuple[str, str]]) -> LegacySpecs:
    all_specs: Dict[str, str] = {}
    known: Dict[str, str] = {}
    unknown: List[str] = []
    for raw_label, raw_value in pairs:
        label = _normalized_spec_label(raw_label)
        value = normalize_text(raw_value).strip(" .،؛")
        if not label or not value:
            continue
        if len(label) > 80 or len(value) > 500:
            continue
        all_specs.setdefault(label, value)
        key = KNOWN_SPEC_LABELS.get(label)
        if key:
            known.setdefault(key, value)
        elif label not in unknown:
            unknown.append(label)
    weight_display = known.get("weight", "")
    return LegacySpecs(
        all_specs=all_specs,
        category=known.get("category", ""),
        weight_display=weight_display,
        weight_grams=parse_weight(weight_display),
        band_type=known.get("band_type", ""),
        stone_color=known.get("stone_color", ""),
        stone_type=known.get("stone_type", ""),
        engraving_type=known.get("engraving_type", ""),
        silver_purity=known.get("silver_purity", ""),
        size=known.get("size", ""),
        unknown_labels=tuple(unknown),
    )


def parse_spec_block_text(text: str) -> LegacySpecs:
    normalized = html.unescape(str(text or ""))
    normalized = normalized.replace("•", "·")
    segments = re.split(r"\s*·\s*|[\r\n]+", normalized)
    pairs: List[Tuple[str, str]] = []
    for segment in segments:
        cleaned = normalize_text(segment)
        if not cleaned:
            continue
        match = re.search(r"[:：]", cleaned)
        if not match:
            continue
        label = cleaned[: match.start()]
        value = cleaned[match.end() :]
        pairs.append((label, value))
    return legacy_specs_from_pairs(pairs)


def extract_legacy_specs(source_html: str) -> LegacySpecs:
    parser = SpecTextHTMLParser()
    parser.feed(source_html)
    candidates: List[LegacySpecs] = []
    plain_text = html.unescape(parser.text)
    for match in re.finditer(r"مشخصات", plain_text):
        chunk = plain_text[match.start() : match.start() + 3000]
        stop = re.search(r"\n\s*(?:توضیحات|نظرات|دیدگاه|محصولات مرتبط)\s*\n", chunk)
        if stop:
            chunk = chunk[: stop.start()]
        candidates.append(parse_spec_block_text(chunk))
    if parser.structured_pairs:
        candidates.append(legacy_specs_from_pairs(parser.structured_pairs))
    if not candidates:
        # Some legacy templates omit the visible section heading but retain the
        # middle-dot field block. Restrict fallback to a known-label window.
        label_match = re.search(
            r"(?:دسته[‌ ]?بندی|وزن|نوع رکاب|نوع سنگ|عیار نقره)\s*[:：]",
            plain_text,
        )
        if label_match:
            candidates.append(
                parse_spec_block_text(
                    plain_text[label_match.start() : label_match.start() + 2500]
                )
            )
    if not candidates:
        return empty_legacy_specs()
    candidates.sort(
        key=lambda item: (
            sum(label in KNOWN_SPEC_LABELS for label in item.all_specs),
            len(item.all_specs),
        ),
        reverse=True,
    )
    best = candidates[0]
    known_count = sum(label in KNOWN_SPEC_LABELS for label in best.all_specs)
    return best if known_count >= 1 else empty_legacy_specs()


def parse_stock(value: Any) -> Tuple[int, Tuple[str, ...]]:
    parsed = parse_decimal(value)
    if parsed is None:
        return 0, ("STOCK_MISSING_OR_INVALID",)
    if parsed < 0:
        return 0, ("STOCK_NEGATIVE_CLAMPED_TO_ZERO",)
    if parsed != parsed.to_integral_value():
        return int(parsed.to_integral_value(rounding=ROUND_FLOOR)), (
            "STOCK_FRACTION_FLOORED",
        )
    return int(parsed), tuple()


def derive_sku(title: str, raw_code: Any, legacy_id: int) -> SKUDecision:
    title_identity = clean_public_product_title(title)
    if title_identity.extracted_code:
        return SKUDecision(
            title_identity.extracted_code, "TITLE_CODE", _raw_trace(raw_code)
        )

    valid_code: Optional[str] = None
    if isinstance(raw_code, bool):
        valid_code = None
    elif isinstance(raw_code, int):
        if 0 < raw_code < 100_000:
            valid_code = str(raw_code)
    elif isinstance(raw_code, Decimal):
        raw_text = format(raw_code, "f")
        if "." not in raw_text and raw_text.isdigit() and int(raw_text) < 100_000:
            valid_code = raw_text
    elif isinstance(raw_code, float):
        # A floating-point cell has already violated the "no decimal point" rule.
        valid_code = None
    else:
        raw_text = normalize_text(raw_code).replace(" ", "")
        if re.fullmatch(r"[0-9]+", raw_text) and 0 < int(raw_text) < 100_000:
            valid_code = raw_text
    if valid_code:
        return SKUDecision(valid_code, "COL27_VALIDATED", _raw_trace(raw_code))
    return SKUDecision(f"NM-{legacy_id}", "FALLBACK_LEGACY_ID", _raw_trace(raw_code))


def classify_stone_from_title(title: str) -> str:
    normalized = normalize_text(title)
    if NO_STONE_RE.search(normalized):
        return "no_stone"
    if LARGE_STONE_RE.search(normalized):
        return "large_stone"
    return "uncertain"


def calculate_pricing(
    *,
    title: str,
    excel_price: Any,
    pre_discount_price: Any,
    weight: Optional[Decimal],
) -> PricingDecision:
    flags: List[str] = []
    current = _positive_toman(excel_price)
    pre_discount = _positive_toman(pre_discount_price)
    stone_class = classify_stone_from_title(title)
    rate = LARGE_STONE_RATE if stone_class == "large_stone" else STANDARD_RATE
    computed: Optional[Decimal] = None
    final: Optional[int] = None
    price_source = "INVALID_REVIEW"

    if current is None:
        flags.append("EXCEL_PRICE_MISSING_OR_INVALID")
    elif weight is None:
        final = round_up_toman(Decimal(current), ROUNDING_STEP)
        price_source = "EXCEL_ONLY"
    else:
        computed = weight * Decimal(rate)
        if Decimal(current) >= computed:
            selected = Decimal(current)
            price_source = "MAX_EXCEL"
        else:
            selected = computed
            price_source = "MAX_CALCULATED"
        final = round_up_toman(selected, ROUNDING_STEP)

    # Luxury pricing policy: COL 10 is trace-only. The storefront exposes one
    # price, so regular price always equals the computed/selected final price.
    regular = final

    return PricingDecision(
        stone_class=stone_class,
        rate_used=rate,
        excel_price_toman=current,
        computed_price_toman=computed,
        final_price_toman=final,
        pre_discount_price_toman=pre_discount,
        regular_price_toman=regular,
        price_source=price_source,
        review_flags=tuple(flags),
    )


def map_category(raw_category: str) -> Tuple[str, Tuple[str, ...]]:
    value = normalize_text(raw_category)
    if "انگشتر" in value:
        return "rings", tuple()
    if "گردنبند" in value or "مدال" in value:
        return "necklaces", tuple()
    if "دستبند" in value:
        return "bracelets", tuple()
    return "rings", ("UNKNOWN_CATEGORY_DEFAULTED_TO_RINGS",)


def _category_display(record: Mapping[str, Any], specs: LegacySpecs) -> str:
    if specs.category:
        return specs.category
    raw = display_text(record.get("category_raw"))
    if raw:
        return raw
    return {
        "rings": "انگشتر نقره",
        "necklaces": "گردنبند و مدال نقره",
        "bracelets": "دستبند نقره",
    }.get(str(record.get("category")), "زیورآلات نقره")


def generate_unique_description(
    record: Mapping[str, Any], specs: LegacySpecs
) -> Tuple[str, str]:
    if not specs.found:
        fallback = display_text(
            record.get("seo_fallback_description")
            or record.get("description")
            or record.get("seo_title")
            or record.get("title")
        )
        return fallback, "SEO_FALLBACK"

    title = display_text(record.get("title"))
    category = _category_display(record, specs)
    legacy_id = int(record.get("legacy_id") or 0)
    sku = display_text(record.get("sku"))
    intros = (
        f"این {category} از مجموعه رادمان سیلور با مشخصات واقعی زیر ارائه می‌شود:",
        f"در این قطعه {category} از رادمان سیلور، جزئیات ساخت و نگین چنین است:",
        f"برای مدل حاضر از گروه {category}، مشخصات ثبت‌شده محصول عبارت است از:",
    )
    bullets: List[str] = []
    if specs.silver_purity:
        bullets.append(f"- جنس: نقره عیار {specs.silver_purity} اصل")
    if specs.stone_type:
        stone = f"- نوع سنگ: {specs.stone_type}"
        if specs.stone_color:
            stone += f"، رنگ {specs.stone_color}"
        bullets.append(stone)
    elif specs.stone_color:
        bullets.append(f"- رنگ نگین: {specs.stone_color}")
    if specs.band_type:
        bullets.append(f"- نوع رکاب: {specs.band_type}")
    if specs.weight_display:
        weight_text = specs.weight_display
        if specs.weight_grams is not None and "گرم" not in weight_text:
            weight_text += " گرم"
        bullets.append(f"- وزن تقریبی: {weight_text}")
    if specs.engraving_type:
        bullets.append(f"- نوع حکاکی: {specs.engraving_type}")
    if specs.size:
        bullets.append(f"- سایز: {specs.size}")
    else:
        bullets.append("- سایز: امکان انتخاب سایز دلخواه هنگام ثبت سفارش")
    for label in specs.unknown_labels:
        value = specs.all_specs.get(label, "")
        if value:
            bullets.append(f"- {label}: {value}")
    if sku:
        bullets.append(f"- کد مدل: {sku}")

    conclusions = (
        "موجودی این قطعه محدود است؛ وزن دقیق ممکن است اندکی با مقدار درج‌شده تفاوت داشته باشد که از ویژگی‌های طبیعی ساخت محصولات نقره است.",
        "این مدل با موجودی محدود عرضه می‌شود. به‌دلیل ماهیت ساخت زیورآلات نقره، اختلاف جزئی وزن با مقدار تقریبی طبیعی است.",
        "تعداد این قطعه محدود است و وزن نهایی می‌تواند در بازه‌ای اندک با عدد ثبت‌شده تفاوت داشته باشد؛ این موضوع در محصولات نقره طبیعی است.",
    )
    description = "\n\n".join(
        [
            title,
            intros[legacy_id % len(intros)],
            "\n".join(bullets),
            conclusions[legacy_id % len(conclusions)],
        ]
    )
    return description, "SPECS_TEMPLATE"


def _apply_pricing_to_record(
    record: Dict[str, Any], weight: Optional[Decimal]
) -> None:
    decision = calculate_pricing(
        title=str(record["title"]),
        excel_price=record.get("excel_price_toman"),
        pre_discount_price=record.get("pre_discount_price_toman"),
        weight=weight,
    )
    record.update(
        {
            "weight_grams": format(weight, "f") if weight is not None else None,
            "stone_class": decision.stone_class,
            "rate_used": decision.rate_used,
            "computed_price_toman": (
                format(decision.computed_price_toman, "f")
                if decision.computed_price_toman is not None
                else None
            ),
            "final_price_toman": decision.final_price_toman,
            "regular_price_toman": decision.regular_price_toman,
            "price_source": decision.price_source,
        }
    )


def apply_specs_to_record(
    source_record: Mapping[str, Any], specs: LegacySpecs
) -> Dict[str, Any]:
    record = dict(source_record)
    flags = list(record.get("review_flags", []))
    record["legacy_specs"] = specs.all_specs
    record["spec_category"] = specs.category
    record["spec_stone_type"] = specs.stone_type
    record["spec_stone_color"] = specs.stone_color
    record["spec_band_type"] = specs.band_type
    record["spec_engraving_type"] = specs.engraving_type
    record["spec_silver_purity"] = specs.silver_purity
    record["spec_size"] = specs.size
    record["spec_weight_grams"] = (
        format(specs.weight_grams, "f") if specs.weight_grams is not None else None
    )
    record["spec_weight_display"] = specs.weight_display
    record["unknown_spec_labels_seen"] = list(specs.unknown_labels)
    record["spec_status"] = "FOUND" if specs.found else "MISSING"

    if specs.category and normalize_text(specs.category) != normalize_text(
        record.get("category_raw")
    ):
        flags.append("SPEC_CATEGORY_MISMATCH")

    excel_weight_raw = record.get("excel_weight_grams")
    if excel_weight_raw in {None, ""}:
        excel_weight_raw = None
    excel_weight = parse_weight(excel_weight_raw)
    live_weight = specs.weight_grams
    previous_final = record.get("final_price_toman")
    if excel_weight is not None:
        record["weight_source"] = "EXCEL"
        record["weight_grams"] = format(excel_weight, "f")
        if live_weight is not None and abs(excel_weight - live_weight) > Decimal("0.5"):
            flags.append("WEIGHT_MISMATCH")
        record["price_reconciled_from_live_weight"] = False
    elif live_weight is not None:
        record["weight_source"] = "LIVE_PAGE"
        _apply_pricing_to_record(record, live_weight)
        record["price_reconciled_from_live_weight"] = True
        flags.append("LIVE_PAGE_WEIGHT_USED_FOR_PRICING")
        if previous_final != record.get("final_price_toman"):
            flags.append("PRICE_RECALCULATED_FROM_LIVE_WEIGHT")
    else:
        record["weight_source"] = "MISSING"
        record["weight_grams"] = None
        record["price_reconciled_from_live_weight"] = False

    description, description_source = generate_unique_description(record, specs)
    record["description"] = description
    record["description_source"] = description_source
    if not specs.found:
        flags.append("SPECS_MISSING_SEO_FALLBACK")
    record["review_flags"] = list(dict.fromkeys(flags))
    record["radman_requires_review"] = "YES" if record["review_flags"] else "NO"
    return record


def _json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {key: _json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [_json_safe(item) for item in value]
    return value


def load_excel_records(
    excel_path: Path,
    *,
    sheet_name: str = SHEET_NAME,
) -> Tuple[List[Dict[str, Any]], List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise ExcelPipelineError(
            "کتابخانه openpyxl نصب نیست؛ اجرا کنید: python3 -m pip install --user openpyxl"
        ) from exc

    path = excel_path.expanduser().resolve()
    if not path.is_file():
        raise ExcelPipelineError(f"فایل Excel پیدا نشد: {path}")
    if "public_html" in str(path):
        raise ExcelPipelineError("خواندن Excel از public_html ممنوع است")
    try:
        workbook = load_workbook(
            path, read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise ExcelPipelineError(f"خواندن Excel ناموفق بود: {exc}") from exc
    warnings: List[str] = []
    records: List[Dict[str, Any]] = []
    try:
        if sheet_name not in workbook.sheetnames:
            raise ExcelPipelineError(
                f"شیت «{sheet_name}» وجود ندارد؛ موجود: {', '.join(workbook.sheetnames)}"
            )
        worksheet = workbook[sheet_name]
        if worksheet.max_column < MIN_REQUIRED_COLUMNS:
            raise ExcelPipelineError(
                f"شیت {worksheet.max_column} ستون دارد؛ حداقل {MIN_REQUIRED_COLUMNS} لازم است"
            )
        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=2, values_only=True), start=2
        ):
            if all(is_blank(value) for value in values):
                continue
            legacy_id = parse_id(_cell(values, COL_ID))
            if legacy_id is None:
                warnings.append(f"ROW_{excel_row}_INVALID_LEGACY_ID")
                continue
            raw_title_value = _cell(values, COL_TITLE)
            raw_title = (
                str(raw_title_value).strip()
                if not is_blank(raw_title_value)
                else f"محصول {legacy_id}"
            )
            title_identity = clean_public_product_title(raw_title)
            title = title_identity.cleaned_title
            raw_category = display_text(_cell(values, COL_CATEGORY))
            availability = normalize_availability(_cell(values, COL_AVAILABILITY))
            active = parse_active(_cell(values, COL_ACTIVE))
            weight_raw = _cell(values, COL_WEIGHT)
            weight = parse_weight(weight_raw)
            raw_code = _cell(values, COL_RAW_CODE)
            sku = derive_sku(raw_title, raw_code, legacy_id)
            stock, stock_flags = parse_stock(_cell(values, COL_STOCK))
            category, category_flags = map_category(raw_category)
            pricing = calculate_pricing(
                title=title,
                excel_price=_cell(values, COL_PRICE),
                pre_discount_price=_cell(values, COL_PRE_DISCOUNT_PRICE),
                weight=weight,
            )
            review_flags = list(stock_flags)
            review_flags.extend(category_flags)
            review_flags.extend(pricing.review_flags)
            review_flags.extend(title_identity.review_flags)
            if not is_blank(weight_raw) and weight is None:
                review_flags.append("WEIGHT_PRESENT_BUT_INVALID")
            description = display_text(_cell(values, COL_SEO_DESCRIPTION))
            seo_title = display_text(_cell(values, COL_SEO_TITLE))
            records.append(
                {
                    "excel_row": excel_row,
                    "legacy_id": legacy_id,
                    "title": title,
                    "legacy_original_title": title_identity.original_title,
                    "old_public_title": "",
                    "new_public_title": title,
                    "title_cleanup_applied": title_identity.cleanup_applied,
                    "extracted_title_code": title_identity.extracted_code,
                    "title_code_label": title_identity.code_label,
                    "legacy_title_cleanup_status": title_identity.cleanup_status,
                    "legacy_title_cleanup_timestamp": datetime.now(TEHRAN).isoformat(),
                    "category_raw": raw_category,
                    "category": category,
                    "excel_price_toman": pricing.excel_price_toman,
                    "pre_discount_price_toman": pricing.pre_discount_price_toman,
                    "excel_weight_grams": (
                        format(weight, "f") if weight is not None else None
                    ),
                    "weight_grams": format(weight, "f") if weight is not None else None,
                    "weight_source": "EXCEL" if weight is not None else "MISSING",
                    "weight_missing": weight is None,
                    "stock": stock,
                    "availability": availability,
                    "active": active,
                    "raw_code": sku.raw_code,
                    "sku": sku.sku,
                    "current_sku": sku.sku,
                    "sku_source": sku.source,
                    "radman_legacy_code": sku.sku,
                    "legacy_identity_key": build_legacy_identity_key(
                        legacy_id, sku.sku
                    ),
                    "sku_title_match": (
                        "NOT_APPLICABLE"
                        if not title_identity.extracted_code
                        else (
                            "YES"
                            if title_identity.extracted_code == sku.sku
                            else "NO"
                        )
                    ),
                    "seo_title": seo_title,
                    "seo_fallback_description": description or seo_title or title,
                    "description": description or seo_title or title,
                    "description_source": "SEO_FALLBACK",
                    "legacy_specs": {},
                    "spec_category": "",
                    "spec_stone_type": "",
                    "spec_stone_color": "",
                    "spec_band_type": "",
                    "spec_engraving_type": "",
                    "spec_silver_purity": "",
                    "spec_size": "",
                    "spec_weight_grams": None,
                    "spec_weight_display": "",
                    "unknown_spec_labels_seen": [],
                    "spec_status": "NOT_FETCHED",
                    "radman_requires_review": (
                        "YES" if review_flags else "NO"
                    ),
                    "price_reconciled_from_live_weight": False,
                    "stone_class": pricing.stone_class,
                    "rate_used": pricing.rate_used,
                    "computed_price_toman": (
                        format(pricing.computed_price_toman, "f")
                        if pricing.computed_price_toman is not None
                        else None
                    ),
                    "final_price_toman": pricing.final_price_toman,
                    "regular_price_toman": pricing.regular_price_toman,
                    "price_source": pricing.price_source,
                    "review_flags": list(dict.fromkeys(review_flags)),
                    "eligible": bool(active is True and availability != "ناموجود"),
                    "legacy_url": "",
                    "image_urls": [],
                    "original_image_paths": [],
                    "selected_import_paths": [],
                    "images_found": 0,
                    "image_status": "NOT_FETCHED",
                    "image_discovery_strategy": "NOT_RUN",
                    "image_qa": None,
                    "action": "UNSELECTED",
                    "wordpress_product_id": None,
                }
            )
    finally:
        workbook.close()
    if not records:
        raise ExcelPipelineError("هیچ محصول معتبر در Excel پیدا نشد")
    return records, warnings


def select_newest(
    records: Sequence[Dict[str, Any]], max_products: int
) -> Tuple[List[Dict[str, Any]], Dict[str, int]]:
    if max_products < 1 or max_products > HARD_MAX_PRODUCTS:
        raise ExcelPipelineError(
            f"MAX_PRODUCTS باید بین 1 و {HARD_MAX_PRODUCTS} باشد"
        )
    ordered = sorted(
        (dict(record) for record in records),
        key=lambda record: (-int(record["legacy_id"]), int(record["excel_row"])),
    )
    inactive = sum(record.get("active") is not True for record in ordered)
    unavailable = sum(
        record.get("active") is True and record.get("availability") == "ناموجود"
        for record in ordered
    )
    eligible = [record for record in ordered if record.get("eligible")]
    selected = eligible[:max_products]
    seen_skus: set[str] = set()
    seen_ids: set[int] = set()
    for record in selected:
        legacy_id = int(record["legacy_id"])
        sku_key = str(record["sku"]).casefold()
        if legacy_id in seen_ids:
            record["action"] = "SKIP_DUPLICATE_BATCH_LEGACY_ID"
            record["review_flags"] = list(
                dict.fromkeys(
                    [*record["review_flags"], "DUPLICATE_BATCH_LEGACY_ID"]
                )
            )
        elif sku_key in seen_skus:
            record["action"] = "SKIP_DUPLICATE_BATCH_SKU"
            record["review_flags"] = list(
                dict.fromkeys([*record["review_flags"], "DUPLICATE_BATCH_SKU"])
            )
        elif record.get("final_price_toman") is None:
            record["action"] = "SKIP_INVALID_PRICE"
        else:
            record["action"] = "PLAN_CREATE_DRAFT"
        seen_ids.add(legacy_id)
        seen_skus.add(sku_key)
    summary = {
        "excel_rows": len(ordered),
        "inactive_skipped": inactive,
        "unavailable_skipped": unavailable,
        "eligible_rows": len(eligible),
        "selected_rows": len(selected),
    }
    return selected, summary


def _canonical_image_url(url: str) -> Optional[str]:
    parsed = urllib.parse.urlsplit(iri_to_uri(url))
    if parsed.scheme not in {"http", "https"} or parsed.hostname != "noghrehmashhad.ir":
        return None
    path = re.sub(
        r"-\d{2,5}x\d{2,5}(?=\.(?:jpe?g|png|webp)$)",
        "",
        parsed.path,
        flags=re.I,
    )
    if not re.search(r"\.(?:jpe?g|png|webp)$", path, re.I):
        return None
    return urllib.parse.urlunsplit(("https", parsed.netloc, path, "", ""))


def extract_gallery_urls(page_url: str, source_html: str) -> List[str]:
    parser = GalleryHTMLParser(page_url)
    parser.feed(source_html)
    targeted: List[str] = []
    fallback: List[str] = []
    seen: set[str] = set()
    for candidate in parser.candidates:
        canonical = _canonical_image_url(candidate)
        if not canonical or canonical in seen:
            continue
        seen.add(canonical)
        path = urllib.parse.urlsplit(canonical).path.casefold()
        basename = Path(path).name.casefold()
        if any(token in basename for token in ("logo", "icon", "avatar", "placeholder")):
            continue
        if any(
            token in path
            for token in (
                "product-images",
                "product_images",
                "/products/",
                "/product/",
                "/uploads/",
            )
        ):
            targeted.append(canonical)
        else:
            fallback.append(canonical)
    if targeted:
        return targeted
    product_marker = re.search(
        r"product[-_ ]gallery|woocommerce-product-gallery|[\"']@type[\"']\s*:\s*[\"']Product[\"']",
        source_html,
        re.I,
    )
    return fallback if product_marker else []


def _links_for_legacy_id(page_url: str, source_html: str, legacy_id: int) -> List[str]:
    discovered = parse_product_links(page_url, source_html)
    hrefs = re.findall(r"href=[\"']([^\"']+)[\"']", source_html, re.I)
    discovered.extend(urllib.parse.urljoin(page_url, html.unescape(href)) for href in hrefs)
    result = []
    for url in discovered:
        parsed = urllib.parse.urlsplit(url)
        if parsed.hostname != "noghrehmashhad.ir":
            continue
        if re.search(rf"/(?:product|p)/{legacy_id}(?:/|$)", parsed.path) or (
            urllib.parse.parse_qs(parsed.query).get("product_id") == [str(legacy_id)]
        ):
            result.append(urllib.parse.urlunsplit(("https", parsed.netloc, parsed.path, parsed.query, "")))
    return list(dict.fromkeys(result))


class LegacyImageDiscovery:
    def __init__(self, fetcher: Optional[RateLimitedFetcher] = None) -> None:
        self.fetcher = fetcher or RateLimitedFetcher(
            user_agent=EXCEL_IMAGE_USER_AGENT
        )
        self._robots_loaded = False
        self._sitemap_by_id: Optional[Dict[int, List[str]]] = None
        self.log: List[Dict[str, Any]] = []

    def load_robots(self) -> None:
        if not self._robots_loaded:
            self.fetcher.load_robots()
            self._robots_loaded = True

    def _try_page(
        self, url: str, strategy: str, legacy_id: int
    ) -> Optional[LegacyPageResult]:
        try:
            source = self.fetcher.fetch_text(url)
        except (OSError, urllib.error.URLError, RuntimeError) as exc:
            self.log.append(
                {
                    "legacy_id": legacy_id,
                    "strategy": strategy,
                    "url": url,
                    "status": "ERROR",
                    "detail": str(exc),
                }
            )
            return None
        images = extract_gallery_urls(url, source)
        specs = extract_legacy_specs(source)
        found = bool(images or specs.found)
        self.log.append(
            {
                "legacy_id": legacy_id,
                "strategy": strategy,
                "url": url,
                "status": "FOUND" if found else "NO_PRODUCT_ASSETS",
                "images": len(images),
                "spec_fields": len(specs.all_specs),
                "unknown_spec_labels": list(specs.unknown_labels),
            }
        )
        if found:
            return LegacyPageResult(url, tuple(images), strategy, specs)
        return None

    def _load_sitemaps(self) -> Dict[int, List[str]]:
        if self._sitemap_by_id is not None:
            return self._sitemap_by_id
        mapping: Dict[int, List[str]] = {}
        queue = [f"{BASE_URL}/sitemap.xml"]
        fetched: set[str] = set()
        while queue and len(fetched) < 40:
            sitemap = queue.pop(0)
            if sitemap in fetched:
                continue
            fetched.add(sitemap)
            try:
                content = self.fetcher.fetch_text(sitemap)
            except (OSError, urllib.error.URLError, RuntimeError):
                continue
            for url in parse_sitemap_urls(content):
                if url.lower().endswith(".xml") or "sitemap" in url.lower():
                    if urllib.parse.urlsplit(url).hostname == "noghrehmashhad.ir":
                        queue.append(url)
                    continue
                match = re.search(r"/(?:product|p)/([0-9]+)(?:/|$)", url)
                if match:
                    mapping.setdefault(int(match.group(1)), []).append(url)
        self._sitemap_by_id = mapping
        return mapping

    def discover(self, legacy_id: int) -> LegacyPageResult:
        self.load_robots()
        direct = (
            f"{BASE_URL}/product/{legacy_id}/",
            f"{BASE_URL}/product/{legacy_id}",
            f"{BASE_URL}/p/{legacy_id}/",
            f"{BASE_URL}/?product_id={legacy_id}",
        )
        for url in direct:
            result = self._try_page(url, "DIRECT_ID_PATTERN", legacy_id)
            if result:
                return result

        for search_url in (
            f"{BASE_URL}/search?q={legacy_id}",
            f"{BASE_URL}/?s={legacy_id}",
        ):
            try:
                search_html = self.fetcher.fetch_text(search_url)
            except (OSError, urllib.error.URLError, RuntimeError) as exc:
                self.log.append(
                    {
                        "legacy_id": legacy_id,
                        "strategy": "SITE_SEARCH",
                        "url": search_url,
                        "status": "ERROR",
                        "detail": str(exc),
                    }
                )
                continue
            for product_url in _links_for_legacy_id(
                search_url, search_html, legacy_id
            ):
                result = self._try_page(product_url, "SITE_SEARCH", legacy_id)
                if result:
                    return result

        for product_url in self._load_sitemaps().get(legacy_id, []):
            result = self._try_page(product_url, "SITEMAP_CACHE", legacy_id)
            if result:
                return result
        self.log.append(
            {
                "legacy_id": legacy_id,
                "strategy": "ALL_STRATEGIES",
                "status": "MISSING",
            }
        )
        return LegacyPageResult("", tuple(), "NOT_FOUND", empty_legacy_specs())


def _image_extension(url: str, content_type: str) -> str:
    suffix = Path(urllib.parse.urlsplit(url).path).suffix.lower()
    if suffix in {".jpg", ".jpeg", ".png", ".webp"}:
        return suffix
    guessed = mimetypes.guess_extension(content_type.split(";", 1)[0].strip())
    return guessed if guessed in {".jpg", ".jpeg", ".png", ".webp"} else ".jpg"


def _sha256_bytes(value: bytes) -> str:
    return hashlib.sha256(value).hexdigest()


def _verify_image_bytes(value: bytes) -> None:
    from PIL import Image

    try:
        with Image.open(io.BytesIO(value)) as image:
            image.verify()
    except Exception as exc:
        raise ExcelPipelineError(f"downloaded gallery bytes are not a valid image: {exc}") from exc


def fetch_images_for_records(
    records: Sequence[Dict[str, Any]],
    *,
    private_dir: Path,
    run_dir: Path,
    discovery: Optional[LegacyImageDiscovery] = None,
) -> List[Dict[str, Any]]:
    service = discovery or LegacyImageDiscovery()
    image_root = private_dir / "legacy-cache" / "original-images"
    qa_dir = run_dir / "image-qa"
    updated: List[Dict[str, Any]] = []
    for source_record in records:
        record = dict(source_record)
        legacy_id = int(record["legacy_id"])
        if str(record.get("action", "")).startswith("SKIP_"):
            record["image_status"] = "SKIPPED"
            record["image_discovery_strategy"] = "NOT_RUN_SKIPPED_PRODUCT"
            updated.append(record)
            print(
                f"[IMAGE] id={legacy_id} strategy=NOT_RUN status=SKIPPED"
            )
            continue
        page = service.discover(legacy_id)
        record = apply_specs_to_record(record, page.specs)
        legacy_url = page.url
        image_urls = list(page.image_urls)
        strategy = page.strategy
        record["legacy_url"] = legacy_url
        record["image_urls"] = image_urls
        record["image_discovery_strategy"] = strategy
        originals: List[str] = []
        downloads: List[Dict[str, Any]] = []
        product_dir = image_root / str(legacy_id)
        product_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
        for position, image_url in enumerate(image_urls, start=1):
            try:
                data, headers = service.fetcher.fetch_bytes(
                    image_url, max_bytes=MAX_IMAGE_BYTES
                )
                content_type = headers.get("content-type", "").lower()
                if content_type and not content_type.startswith("image/"):
                    raise ExcelPipelineError("gallery URL returned non-image content")
                _verify_image_bytes(data)
                digest = _sha256_bytes(data)
                extension = _image_extension(image_url, content_type)
                destination = product_dir / f"{position:02d}-original{extension}"
                if destination.exists() and hashlib.sha256(
                    destination.read_bytes()
                ).hexdigest() != digest:
                    destination = product_dir / (
                        f"{position:02d}-original-{digest[:12]}{extension}"
                    )
                if not destination.exists():
                    temporary = destination.with_suffix(destination.suffix + ".tmp")
                    temporary.write_bytes(data)
                    os.chmod(temporary, 0o600)
                    os.replace(temporary, destination)
                originals.append(str(destination))
                downloads.append(
                    {
                        "position": position,
                        "source_url": image_url,
                        "local_path": str(destination),
                        "sha256": digest,
                        "bytes": len(data),
                    }
                )
            except (OSError, urllib.error.URLError, RuntimeError) as exc:
                downloads.append(
                    {
                        "position": position,
                        "source_url": image_url,
                        "error": str(exc),
                    }
                )
        record["original_image_paths"] = originals
        record["downloaded_images"] = downloads
        record["images_found"] = len(originals)
        if any(item.get("error") for item in downloads):
            record["review_flags"] = list(
                dict.fromkeys(
                    [*record.get("review_flags", []), "IMAGE_DOWNLOAD_ERROR"]
                )
            )
        if originals:
            qa = process_product(
                str(legacy_id),
                [Path(path) for path in originals],
                private_dir,
                qa_dir,
            )
            record["image_qa"] = qa
            record["selected_import_paths"] = qa.get(
                "selected_import_paths", originals
            )
            record["image_status"] = "READY"
        else:
            record["image_qa"] = None
            record["selected_import_paths"] = []
            record["image_status"] = "MISSING"
            record["review_flags"] = list(
                dict.fromkeys([*record.get("review_flags", []), "IMAGES_MISSING"])
            )
        if record.get("review_flags"):
            record["radman_requires_review"] = "YES"
        if record.get("action") == "PLAN_CREATE_DRAFT":
            record["action"] = "READY_DRAFT"
        updated.append(record)
        print(
            f"[IMAGE] id={legacy_id} strategy={strategy} "
            f"found={len(originals)} status={record['image_status']}"
        )
    write_json_atomic(run_dir / "image-discovery-log.json", service.log)
    return updated


def fetch_specs_for_records(
    records: Sequence[Dict[str, Any]],
    *,
    run_dir: Path,
    discovery: Optional[LegacyImageDiscovery] = None,
) -> List[Dict[str, Any]]:
    service = discovery or LegacyImageDiscovery()
    updated: List[Dict[str, Any]] = []
    for source_record in records:
        legacy_id = int(source_record["legacy_id"])
        page = service.discover(legacy_id)
        record = apply_specs_to_record(source_record, page.specs)
        record["legacy_url"] = page.url or str(source_record.get("legacy_url") or "")
        record["image_discovery_strategy"] = page.strategy
        record["action"] = "ENRICH_READY"
        updated.append(record)
        print(
            f"[SPECS] id={legacy_id} strategy={page.strategy} "
            f"fields={len(page.specs.all_specs)} description={record['description_source']}"
        )
    write_json_atomic(run_dir / "spec-discovery-log.json", service.log)
    return updated


def _b64(value: str) -> str:
    return base64.b64encode(value.encode("utf-8")).decode("ascii")


class ExcelDraftGateway(WPGateway):
    def get_currency(self) -> str:
        return self.eval_scalar("echo (string) get_option('woocommerce_currency', '');")

    def find_by_legacy_id(self, legacy_id: str) -> Optional[int]:
        encoded = _b64(legacy_id)
        php = f"""
$v=base64_decode('{encoded}');
$q=new WP_Query(array('post_type'=>'product','post_status'=>'any','posts_per_page'=>2,'fields'=>'ids',
 'meta_query'=>array('relation'=>'OR',
  array('key'=>'legacy_product_id','value'=>$v,'compare'=>'='),
  array('key'=>'radman_legacy_id','value'=>$v,'compare'=>'='),
  array('key'=>'_legacy_store_id','value'=>$v,'compare'=>'=')
 )));
if (count($q->posts)>1) {{ fwrite(STDERR, 'duplicate legacy product ID metadata'); exit(8); }}
if ($q->posts) {{ echo (string)$q->posts[0]; }}
"""
        raw = self.eval_scalar(php)
        return int(raw) if raw.isdigit() and int(raw) > 0 else None

    def list_draft_legacy_products(self, limit: int) -> List[Dict[str, Any]]:
        php = f"""
$q=new WP_Query(array('post_type'=>'product','post_status'=>'draft','posts_per_page'=>{int(limit)},
 'meta_key'=>'legacy_product_id','orderby'=>'meta_value_num','order'=>'DESC','fields'=>'ids'));
$out=array();
foreach ($q->posts as $id) {{
 $p=wc_get_product($id);
 if (!$p) {{ continue; }}
 $out[]=array(
  'product_id'=>(int)$id,
  'legacy_id'=>(string)get_post_meta($id,'legacy_product_id',true),
  'public_title'=>(string)$p->get_name('edit'),
  'sku'=>(string)$p->get_sku('edit'),
  'legacy_raw_code'=>(string)get_post_meta($id,'legacy_raw_code',true),
  'legacy_url'=>(string)get_post_meta($id,'legacy_url',true),
  'legacy_original_title'=>(string)get_post_meta($id,'legacy_original_title',true),
  'legacy_identity_key'=>(string)get_post_meta($id,'legacy_identity_key',true),
  'title_cleanup_status'=>(string)get_post_meta($id,'legacy_title_cleanup_status',true),
  'title_cleanup_timestamp'=>(string)get_post_meta($id,'legacy_title_cleanup_timestamp',true),
  'review_flags'=>(string)get_post_meta($id,'radman_review_flags',true),
  'price'=>(string)$p->get_price('edit'),
  'regular_price'=>(string)$p->get_regular_price('edit')
 );
}}
echo wp_json_encode($out);
"""
        result = self.eval_json(php)
        if not isinstance(result, list):
            raise WPCliError("invalid existing Draft list")
        return [dict(item) for item in result if isinstance(item, dict)]

    def enrich_existing_draft(
        self,
        record: Mapping[str, Any],
        product_id: int,
        *,
        update_price: bool,
    ) -> Dict[str, Any]:
        meta = {
            "legacy_product_id": str(record["legacy_id"]),
            "_legacy_store_id": str(record["legacy_id"]),
            "legacy_raw_code": str(record.get("raw_code") or ""),
            "legacy_url": str(record.get("legacy_url") or ""),
            "radman_legacy_code": str(record.get("radman_legacy_code") or record["sku"]),
            "legacy_original_title": str(record.get("legacy_original_title") or record["title"]),
            "legacy_identity_key": str(record.get("legacy_identity_key") or ""),
            "legacy_title_cleanup_status": str(
                record.get("legacy_title_cleanup_status") or "UNCHANGED"
            ),
            "legacy_title_cleanup_timestamp": str(
                record.get("legacy_title_cleanup_timestamp") or ""
            ),
            "radman_legacy_specs": json.dumps(
                record.get("legacy_specs", {}), ensure_ascii=False
            ),
            "radman_spec_stone_type": str(record.get("spec_stone_type") or ""),
            "radman_spec_stone_color": str(record.get("spec_stone_color") or ""),
            "radman_spec_band_type": str(record.get("spec_band_type") or ""),
            "radman_spec_engraving_type": str(
                record.get("spec_engraving_type") or ""
            ),
            "radman_spec_silver_purity": str(
                record.get("spec_silver_purity") or ""
            ),
            "radman_spec_size": str(record.get("spec_size") or ""),
            "radman_spec_weight_grams": str(
                record.get("spec_weight_grams") or ""
            ),
            "radman_spec_weight_display": str(
                record.get("spec_weight_display") or ""
            ),
            "weight_source": str(record.get("weight_source") or "MISSING"),
            "description_source": str(
                record.get("description_source") or "SEO_FALLBACK"
            ),
            "radman_requires_review": str(
                record.get("radman_requires_review") or "NO"
            ),
            "unknown_spec_labels_seen": ",".join(
                record.get("unknown_spec_labels_seen", [])
            ),
            "weight_grams": str(record.get("weight_grams") or ""),
            "silver_weight_grams": str(record.get("weight_grams") or ""),
            "price_source": str(record.get("price_source") or ""),
            "rate_used": str(record.get("rate_used") or ""),
            "computed_price": str(record.get("computed_price_toman") or ""),
            "final_price": str(record.get("final_price_toman") or ""),
            "radman_review_flags": " | ".join(record.get("review_flags", [])),
            "radman_import_source": "excel_1000_pipeline",
            "radman_import_version": PIPELINE_VERSION,
        }
        payload = {
            "product_id": int(product_id),
            "legacy_id": str(record["legacy_id"]),
            "public_title": str(record.get("title") or ""),
            "description": str(record.get("description") or ""),
            "short_description": (
                str(record.get("seo_fallback_description") or "")
                if record.get("description_source") == "SEO_FALLBACK"
                else ""
            ),
            "update_price": bool(update_price),
            "regular_price": record.get("regular_price_toman"),
            "meta": meta,
        }
        encoded = _b64(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        )
        php = f"""
$d=json_decode(base64_decode('{encoded}'), true);
$p=wc_get_product((int)$d['product_id']);
if (!$p || $p->get_status('edit') !== 'draft') {{ fwrite(STDERR, 'Draft product not found'); exit(12); }}
$legacy=(string)$p->get_meta('legacy_product_id', true, 'edit');
if ($legacy !== (string)$d['legacy_id']) {{ fwrite(STDERR, 'legacy ID mismatch'); exit(13); }}
$p->set_name($d['public_title']);
$p->set_description($d['description']);
$p->set_short_description($d['short_description']);
if ($d['update_price']) {{
 $p->set_regular_price((string)$d['regular_price']);
}}
$luxury_regular=(string)$p->get_regular_price('edit');
$p->set_price($luxury_regular);
foreach ($d['meta'] as $key=>$value) {{ $p->update_meta_data($key, (string)$value); }}
$id=$p->save();
delete_post_meta($id, '_sale_price');
update_post_meta($id, '_price', $luxury_regular);
wc_delete_product_transients($id);
echo wp_json_encode(array('id'=>(int)$id,'status'=>$p->get_status('edit'),'price_updated'=>(bool)$d['update_price']));
"""
        result = self.eval_json(php)
        if not isinstance(result, dict) or result.get("status") != "draft":
            raise WPCliError(f"invalid enrichment result for {record['legacy_id']}")
        return dict(result)

    def create_excel_draft(self, record: Mapping[str, Any], category_id: int) -> int:
        metadata = {
            "legacy_product_id": str(record["legacy_id"]),
            "_legacy_store_id": str(record["legacy_id"]),
            "legacy_raw_code": str(record.get("raw_code") or ""),
            "legacy_url": str(record.get("legacy_url") or ""),
            "radman_legacy_code": str(record.get("radman_legacy_code") or record["sku"]),
            "legacy_original_title": str(record.get("legacy_original_title") or record["title"]),
            "legacy_identity_key": str(record.get("legacy_identity_key") or ""),
            "legacy_title_cleanup_status": str(
                record.get("legacy_title_cleanup_status") or "UNCHANGED"
            ),
            "legacy_title_cleanup_timestamp": str(
                record.get("legacy_title_cleanup_timestamp") or ""
            ),
            "radman_legacy_specs": json.dumps(
                record.get("legacy_specs", {}), ensure_ascii=False
            ),
            "radman_spec_stone_type": str(record.get("spec_stone_type") or ""),
            "radman_spec_stone_color": str(record.get("spec_stone_color") or ""),
            "radman_spec_band_type": str(record.get("spec_band_type") or ""),
            "radman_spec_engraving_type": str(
                record.get("spec_engraving_type") or ""
            ),
            "radman_spec_silver_purity": str(
                record.get("spec_silver_purity") or ""
            ),
            "radman_spec_size": str(record.get("spec_size") or ""),
            "radman_spec_weight_grams": str(
                record.get("spec_weight_grams") or ""
            ),
            "radman_spec_weight_display": str(
                record.get("spec_weight_display") or ""
            ),
            "weight_source": str(record.get("weight_source") or "MISSING"),
            "description_source": str(
                record.get("description_source") or "SEO_FALLBACK"
            ),
            "radman_requires_review": str(
                record.get("radman_requires_review") or "NO"
            ),
            "unknown_spec_labels_seen": ",".join(
                record.get("unknown_spec_labels_seen", [])
            ),
            "weight_grams": str(record.get("weight_grams") or ""),
            "silver_weight_grams": str(record.get("weight_grams") or ""),
            "price_source": str(record["price_source"]),
            "rate_used": str(record["rate_used"]),
            "computed_price": str(record.get("computed_price_toman") or ""),
            "final_price": str(record["final_price_toman"]),
            "excel_price_toman": str(record["excel_price_toman"]),
            "pre_discount_price_toman": str(
                record.get("pre_discount_price_toman") or ""
            ),
            "stone_class": str(record["stone_class"]),
            "image_status": str(record["image_status"]),
            "image_discovery_strategy": str(
                record.get("image_discovery_strategy") or ""
            ),
            "excel_category_raw": str(record.get("category_raw") or ""),
            "excel_row": str(record["excel_row"]),
            "radman_import_source": "excel_1000_pipeline",
            "radman_import_version": PIPELINE_VERSION,
            "radman_review_flags": " | ".join(record.get("review_flags", [])),
            "pricing_mode": "manual_locked",
            "manual_price_toman": str(record["final_price_toman"]),
            "price_locked": "1",
            "rounding_step_toman": str(ROUNDING_STEP),
        }
        payload = {
            "sku": record["sku"],
            "name": record["title"],
            "description": record["description"],
            "short_description": (
                str(record.get("seo_fallback_description") or "")
                if record.get("description_source") == "SEO_FALLBACK"
                else ""
            ),
            "category_id": int(category_id),
            "regular_price": int(record["regular_price_toman"]),
            "stock": int(record["stock"]),
            "meta": metadata,
        }
        encoded = _b64(
            json.dumps(payload, ensure_ascii=False, separators=(",", ":"))
        )
        php = f"""
if (!function_exists('wc_get_product_id_by_sku')) {{ fwrite(STDERR, 'WooCommerce unavailable'); exit(3); }}
$d=json_decode(base64_decode('{encoded}'), true);
if (wc_get_product_id_by_sku($d['sku'])) {{ fwrite(STDERR, 'SKU conflict during create'); exit(9); }}
$legacy=(string)$d['meta']['legacy_product_id'];
$q=new WP_Query(array('post_type'=>'product','post_status'=>'any','posts_per_page'=>1,'fields'=>'ids',
 'meta_query'=>array('relation'=>'OR',
  array('key'=>'legacy_product_id','value'=>$legacy,'compare'=>'='),
  array('key'=>'radman_legacy_id','value'=>$legacy,'compare'=>'='),
  array('key'=>'_legacy_store_id','value'=>$legacy,'compare'=>'=')
 )));
if ($q->posts) {{ fwrite(STDERR, 'legacy ID appeared during create'); exit(10); }}
$p=new WC_Product_Simple();
$p->set_sku($d['sku']);
$p->set_status('draft');
$p->set_catalog_visibility('visible');
$p->set_name($d['name']);
$p->set_description($d['description']);
$p->set_short_description($d['short_description']);
$p->set_regular_price((string)$d['regular_price']);
$p->set_price((string)$d['regular_price']);
$p->set_manage_stock(true);
$p->set_stock_quantity((int)$d['stock']);
$p->set_stock_status(((int)$d['stock']) > 0 ? 'instock' : 'outofstock');
$p->set_backorders('no');
$p->set_category_ids(array((int)$d['category_id']));
foreach ($d['meta'] as $key=>$value) {{
 $p->update_meta_data($key, (string)$value);
}}
$id=$p->save();
if (!$id || $p->get_status('edit') !== 'draft' || $p->get_backorders('edit') !== 'no') {{
 fwrite(STDERR, 'Draft/backorder verification failed'); exit(11);
}}
echo wp_json_encode(array('id'=>(int)$id,'status'=>$p->get_status('edit')));
"""
        result = self.eval_json(php)
        if not isinstance(result, dict) or result.get("status") != "draft":
            raise WPCliError(f"invalid Draft creation result for {record['sku']}")
        return int(result["id"])


def require_readonly_wp_environment(wp_path: str) -> None:
    errors = []
    if os.environ.get("APP_ENV") != EXPECTED_APP_ENV:
        errors.append("APP_ENV must equal staging")
    if os.environ.get("WP_URL") != EXPECTED_WP_URL:
        errors.append(f"WP_URL must equal {EXPECTED_WP_URL}")
    if wp_path != EXPECTED_WP_PATH:
        errors.append(f"WP_PATH must equal {EXPECTED_WP_PATH}")
    if "public_html" in wp_path:
        errors.append("public_html is prohibited")
    if errors:
        raise ExcelPipelineError("; ".join(errors))


def require_import_environment(wp_path: str) -> Path:
    require_readonly_wp_environment(wp_path)
    errors = []
    if os.environ.get("CONFIRM_STAGING_APPLY") != "YES":
        errors.append("CONFIRM_STAGING_APPLY must equal YES")
    backup = Path(os.environ.get("RADMAN_DB_BACKUP_PATH", ""))
    if not backup.is_file() or backup.stat().st_size <= 0:
        errors.append("RADMAN_DB_BACKUP_PATH must be a non-empty database backup")
    if "public_html" in str(backup):
        errors.append("database backup cannot be under public_html")
    if errors:
        raise ExcelPipelineError("; ".join(errors))
    return backup


def preflight_import(
    records: Sequence[Dict[str, Any]], gateway: ExcelDraftGateway
) -> List[Tuple[Dict[str, Any], str, Optional[int]]]:
    if gateway.get_currency() != REQUIRED_CURRENCY:
        raise ExcelPipelineError("WooCommerce currency must be IRT")
    decisions = []
    seen_skus: set[str] = set()
    seen_ids: set[str] = set()
    for record in records:
        legacy_id = str(record["legacy_id"])
        sku_key = str(record["sku"]).casefold()
        if record.get("action") in {
            "SKIP_DUPLICATE_BATCH_SKU",
            "SKIP_DUPLICATE_BATCH_LEGACY_ID",
            "SKIP_INVALID_PRICE",
        }:
            decisions.append((record, str(record["action"]), None))
            continue
        if legacy_id in seen_ids:
            decisions.append((record, "SKIP_DUPLICATE_BATCH_LEGACY_ID", None))
            continue
        if sku_key in seen_skus:
            decisions.append((record, "SKIP_DUPLICATE_BATCH_SKU", None))
            continue
        seen_ids.add(legacy_id)
        seen_skus.add(sku_key)
        existing_legacy = gateway.find_by_legacy_id(legacy_id)
        if existing_legacy:
            decisions.append((record, "SKIP_EXISTING_LEGACY_ID", existing_legacy))
            continue
        existing_sku = gateway.find_product_id(str(record["sku"]))
        if existing_sku:
            record["review_flags"] = list(
                dict.fromkeys(
                    [*record.get("review_flags", []), "WORDPRESS_SKU_CONFLICT"]
                )
            )
            decisions.append((record, "SKIP_SKU_CONFLICT", existing_sku))
            continue
        decisions.append((record, "CREATE_DRAFT", None))
    return decisions


def import_records(
    records: Sequence[Dict[str, Any]], gateway: ExcelDraftGateway
) -> List[Dict[str, Any]]:
    decisions = preflight_import(records, gateway)
    actions: List[Dict[str, Any]] = []
    for record, decision, existing_id in decisions:
        if decision != "CREATE_DRAFT":
            record["action"] = decision
            record["wordpress_product_id"] = existing_id
            actions.append(
                {
                    "legacy_id": record["legacy_id"],
                    "sku": record["sku"],
                    "action": decision,
                    "product_id": existing_id,
                }
            )
            continue
        record["radman_requires_review"] = (
            "YES" if record.get("review_flags") else "NO"
        )
        valid_media: List[Path] = []
        for raw_path in record.get("selected_import_paths", []):
            path = Path(raw_path)
            if path.is_file():
                valid_media.append(path)
            else:
                record["review_flags"] = list(
                    dict.fromkeys(
                        [*record.get("review_flags", []), "SELECTED_IMAGE_MISSING"]
                    )
                )
        if record.get("image_status") == "READY" and not valid_media:
            record["image_status"] = "MISSING"
        category_id = gateway.resolve_category_id(str(record["category"]))
        product_id = gateway.create_excel_draft(record, category_id)
        attachment_ids: List[int] = []
        for path in valid_media:
            attachment_ids.append(
                gateway.import_image(
                    path,
                    str(record["title"]),
                    product_id,
                    str(record["sku"]),
                )
            )
        if attachment_ids:
            gateway.set_product_images(product_id, attachment_ids)
        record["action"] = "CREATED_DRAFT"
        record["wordpress_product_id"] = product_id
        record["attachment_ids"] = attachment_ids
        actions.append(
            {
                "legacy_id": record["legacy_id"],
                "sku": record["sku"],
                "action": "CREATED_DRAFT",
                "product_id": product_id,
                "attachment_ids": attachment_ids,
                "image_status": record.get("image_status"),
            }
        )
    return actions


def enrich_existing_records(
    records: Sequence[Dict[str, Any]],
    gateway: ExcelDraftGateway,
) -> List[Dict[str, Any]]:
    actions: List[Dict[str, Any]] = []
    for record in records:
        product_id = int(record["wordpress_product_id"])
        current_price = _positive_toman(record.get("wordpress_current_price"))
        final_price = record.get("final_price_toman")
        update_price = bool(
            record.get("price_reconciled_from_live_weight")
            and final_price is not None
            and current_price != int(final_price)
        )
        result = gateway.enrich_existing_draft(
            record,
            product_id,
            update_price=update_price,
        )
        record["action"] = "ENRICHED_DRAFT"
        record["description_updated"] = "YES"
        record["specs_found_count"] = len(record.get("legacy_specs", {}))
        record["price_updated_during_enrichment"] = bool(
            result.get("price_updated")
        )
        record["price_changed"] = bool(result.get("price_updated"))
        actions.append(
            {
                "legacy_id": record["legacy_id"],
                "sku": record["sku"],
                "product_id": product_id,
                "old_public_title": record.get("old_public_title"),
                "new_public_title": record.get("title"),
                "title_cleanup_applied": record.get("title_cleanup_applied"),
                "sku_title_match": record.get("sku_title_match"),
                "action": "ENRICHED_DRAFT",
                "description_source": record.get("description_source"),
                "weight_source": record.get("weight_source"),
                "description_updated": "YES",
                "specs_found_count": len(record.get("legacy_specs", {})),
                "price_updated": bool(result.get("price_updated")),
                "review_flags": list(record.get("review_flags", [])),
            }
        )
    return actions


def prepare_existing_enrichment(
    all_excel_records: Sequence[Dict[str, Any]],
    gateway: ExcelDraftGateway,
    max_products: int,
) -> Tuple[List[Dict[str, Any]], List[Dict[str, Any]]]:
    existing = gateway.list_draft_legacy_products(max_products)
    by_legacy_id = {
        str(record["legacy_id"]): dict(record) for record in all_excel_records
    }
    matched: List[Dict[str, Any]] = []
    missing: List[Dict[str, Any]] = []
    for item in existing:
        legacy_id = str(item.get("legacy_id") or "")
        excel_record = by_legacy_id.get(legacy_id)
        if excel_record is None:
            missing.append(
                {
                    "legacy_id": legacy_id,
                    "wordpress_product_id": item.get("product_id"),
                    "action": "SKIP_LEGACY_ID_NOT_IN_EXCEL",
                }
            )
            continue
        existing_sku = str(item.get("sku") or excel_record["sku"])
        original_title = str(
            item.get("legacy_original_title")
            or excel_record.get("legacy_original_title")
            or item.get("public_title")
            or excel_record["title"]
        )
        title_identity = clean_public_product_title(original_title)
        old_public_title = str(item.get("public_title") or excel_record["title"])
        normalized_existing_sku = normalize_identity_digits(existing_sku).strip()
        flags = list(excel_record.get("review_flags", []))
        flags.extend(title_identity.review_flags)
        if (
            title_identity.extracted_code
            and title_identity.extracted_code != normalized_existing_sku
        ):
            flags.append("SKU_TITLE_MISMATCH")
            sku_title_match = "NO"
        elif title_identity.extracted_code:
            sku_title_match = "YES"
        else:
            sku_title_match = "NOT_APPLICABLE"
        cleanup_status = title_identity.cleanup_status
        if "SKU_TITLE_MISMATCH" in flags or title_identity.review_flags:
            cleanup_status = "REVIEW"

        excel_record["wordpress_product_id"] = int(item["product_id"])
        excel_record["wordpress_current_price"] = item.get("price")
        excel_record["wordpress_regular_price"] = item.get("regular_price")
        excel_record["old_public_title"] = old_public_title
        excel_record["new_public_title"] = title_identity.cleaned_title
        excel_record["title"] = title_identity.cleaned_title
        excel_record["title_cleanup_applied"] = (
            normalize_text(old_public_title)
            != normalize_text(title_identity.cleaned_title)
        )
        excel_record["extracted_title_code"] = title_identity.extracted_code
        excel_record["title_code_label"] = title_identity.code_label
        excel_record["legacy_original_title"] = original_title
        excel_record["legacy_title_cleanup_status"] = cleanup_status
        excel_record["legacy_title_cleanup_timestamp"] = str(
            item.get("title_cleanup_timestamp")
            or excel_record.get("legacy_title_cleanup_timestamp")
            or datetime.now(TEHRAN).isoformat()
        )
        excel_record["legacy_url"] = str(
            item.get("legacy_url") or excel_record.get("legacy_url") or ""
        )
        excel_record["raw_code"] = str(
            item.get("legacy_raw_code") or excel_record.get("raw_code") or ""
        )
        excel_record["sku"] = existing_sku
        excel_record["current_sku"] = existing_sku
        excel_record["radman_legacy_code"] = normalized_existing_sku
        excel_record["sku_title_match"] = sku_title_match
        excel_record["legacy_identity_key"] = build_legacy_identity_key(
            legacy_id, normalized_existing_sku
        )
        excel_record["review_flags"] = list(dict.fromkeys(flags))
        excel_record["radman_requires_review"] = (
            "YES" if excel_record["review_flags"] else "NO"
        )
        matched.append(excel_record)
    return matched, missing


def now_slug() -> str:
    return datetime.now(TEHRAN).strftime("%Y%m%dT%H%M%S%f%z")


def write_json_atomic(path: Path, value: Any) -> None:
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(
        json.dumps(_json_safe(value), ensure_ascii=False, indent=2),
        encoding="utf-8",
    )
    os.chmod(temporary, 0o600)
    os.replace(temporary, path)


def _report_row(record: Mapping[str, Any]) -> Dict[str, Any]:
    return {
        "wp_id": record.get("wordpress_product_id") or "",
        "old_public_title": record.get("old_public_title", ""),
        "new_public_title": record.get("new_public_title") or record.get("title", ""),
        "title_cleanup_applied": record.get("title_cleanup_applied", False),
        "extracted_title_code": record.get("extracted_title_code", ""),
        "current_sku": record.get("current_sku") or record.get("sku", ""),
        "sku_title_match": record.get("sku_title_match", "NOT_APPLICABLE"),
        "legacy_product_id": record.get("legacy_id", ""),
        "legacy_raw_code": record.get("raw_code", ""),
        "legacy_url": record.get("legacy_url", ""),
        "identity_key": record.get("legacy_identity_key", ""),
        "title_cleanup_status": record.get("legacy_title_cleanup_status", ""),
        "description_updated": record.get("description_updated", ""),
        "specs_found_count": record.get(
            "specs_found_count", len(record.get("legacy_specs", {}))
        ),
        "price_changed": record.get("price_changed", False),
        "legacy_id": record.get("legacy_id", ""),
        "sku": record.get("sku", ""),
        "sku_source": record.get("sku_source", ""),
        "title": record.get("title", ""),
        "category_raw": record.get("category_raw", ""),
        "category": record.get("category", ""),
        "weight_grams": record.get("weight_grams") or "",
        "excel_price_toman": record.get("excel_price_toman") or "",
        "pre_discount_price_toman": record.get("pre_discount_price_toman") or "",
        "computed_price_toman": record.get("computed_price_toman") or "",
        "final_price_toman": record.get("final_price_toman") or "",
        "regular_price_toman": record.get("regular_price_toman") or "",
        "price_source": record.get("price_source", ""),
        "rate_used": record.get("rate_used", ""),
        "stone_class": record.get("stone_class", ""),
        "stone_type": record.get("spec_stone_type", ""),
        "stone_color": record.get("spec_stone_color", ""),
        "band_type": record.get("spec_band_type", ""),
        "silver_purity": record.get("spec_silver_purity", ""),
        "spec_weight": record.get("spec_weight_grams") or "",
        "weight_source": record.get("weight_source", "MISSING"),
        "description_source": record.get("description_source", "SEO_FALLBACK"),
        "unknown_spec_labels_seen": ",".join(
            record.get("unknown_spec_labels_seen", [])
        ),
        "stock": record.get("stock", 0),
        "images_found": record.get("images_found", 0),
        "image_status": record.get("image_status", ""),
        "image_discovery_strategy": record.get("image_discovery_strategy", ""),
        "action": record.get("action", ""),
        "review_flags": " | ".join(record.get("review_flags", [])),
        "wordpress_product_id": record.get("wordpress_product_id") or "",
    }


def write_reports(
    records: Sequence[Mapping[str, Any]],
    run_dir: Path,
    run_stamp: str,
    selection_summary: Mapping[str, int],
) -> Tuple[Path, Path]:
    csv_path = run_dir / f"excel-import-{run_stamp}.csv"
    txt_path = run_dir / f"excel-import-{run_stamp}-fa.txt"
    run_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
    csv_temporary = csv_path.with_suffix(csv_path.suffix + ".tmp")
    with csv_temporary.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=REPORT_COLUMNS)
        writer.writeheader()
        for record in records:
            writer.writerow(_report_row(record))
    os.chmod(csv_temporary, 0o600)
    os.replace(csv_temporary, csv_path)

    missing_images = sum(record.get("image_status") == "MISSING" for record in records)
    review = sum(bool(record.get("review_flags")) for record in records)
    specs_template_count = sum(
        record.get("description_source") == "SPECS_TEMPLATE" for record in records
    )
    unknown_labels = Counter(
        label
        for record in records
        for label in record.get("unknown_spec_labels_seen", [])
    )
    lines = [
        "گزارش خط لوله Excel برای جدیدترین محصولات رادمان",
        f"زمان: {datetime.now(TEHRAN).isoformat(timespec='seconds')}",
        "منبع داده محصول: فقط Excel (قیمت‌ها تومان هستند)",
        "ترتیب انتخاب: شناسه قدیمی نزولی؛ ID بالاتر جدیدتر است",
        f"ردیف Excel: {selection_summary.get('excel_rows', 0)}",
        f"فعالِ قابل انتخاب: {selection_summary.get('eligible_rows', 0)}",
        f"انتخاب‌شده: {selection_summary.get('selected_rows', len(records))}",
        f"غیرفعال ردشده: {selection_summary.get('inactive_skipped', 0)}",
        f"ناموجود ردشده: {selection_summary.get('unavailable_skipped', 0)}",
        f"بدون تصویر: {missing_images}",
        f"توضیح ساخته‌شده از مشخصات واقعی: {specs_template_count}",
        f"نیازمند بازبینی: {review}",
        "برچسب‌های مشخصات ناشناخته در batch: "
        + (
            "، ".join(f"{label} ({count})" for label, count in unknown_labels.items())
            if unknown_labels
            else "ندارد"
        ),
        "",
        "wp_id | عنوان قبلی -> عنوان جدید | پاکسازی | کد عنوان | SKU/تطبیق | legacy_id/identity | مشخصات/توضیح | قیمت تغییر کرد | اقدام | بازبینی",
    ]
    for record in records:
        lines.append(
            "{wp_id} | {old_title} -> {new_title} | {cleaned} | {code} | "
            "{sku}/{match} | {legacy_id}/{identity} | {spec_count}/{description} | "
            "{price_changed} | {action} | {flags}".format(
                wp_id=record.get("wordpress_product_id") or "—",
                old_title=str(record.get("old_public_title") or "—").replace("|", "/"),
                new_title=str(record.get("title") or "—").replace("|", "/"),
                cleaned="بله" if record.get("title_cleanup_applied") else "خیر",
                code=record.get("extracted_title_code") or "—",
                sku=record.get("sku") or "—",
                match=record.get("sku_title_match") or "NOT_APPLICABLE",
                legacy_id=record.get("legacy_id"),
                identity=record.get("legacy_identity_key") or "—",
                spec_count=len(record.get("legacy_specs", {})),
                description=record.get("description_source") or "SEO_FALLBACK",
                price_changed="بله" if record.get("price_changed") else "خیر",
                action=record.get("action"),
                flags=",".join(record.get("review_flags", [])) or "—",
            )
        )
    txt_temporary = txt_path.with_suffix(txt_path.suffix + ".tmp")
    txt_temporary.write_text("\n".join(lines) + "\n", encoding="utf-8")
    os.chmod(txt_temporary, 0o600)
    os.replace(txt_temporary, txt_path)
    return csv_path, txt_path


def write_identity_report(
    rows: Sequence[Mapping[str, Any]], private_dir: Path
) -> Tuple[Path, Path]:
    stamp = now_slug()
    run_dir = (
        private_dir / "legacy-cache" / "runs" / f"identity-report-{stamp}"
    )
    run_dir.mkdir(parents=True, exist_ok=True, mode=0o700)
    columns = (
        "wp_id",
        "public_title",
        "sku",
        "legacy_product_id",
        "legacy_raw_code",
        "legacy_url",
        "identity_key",
        "title_cleanup_status",
        "review_flags",
    )
    csv_path = run_dir / f"identity-report-{stamp}.csv"
    txt_path = run_dir / f"identity-report-{stamp}-fa.txt"
    normalized = []
    for item in rows:
        sku = str(item.get("sku") or "")
        legacy_id = str(item.get("legacy_id") or "")
        identity_key = str(item.get("legacy_identity_key") or "")
        if not identity_key and sku and legacy_id:
            identity_key = build_legacy_identity_key(legacy_id, sku)
        normalized.append(
            {
                "wp_id": item.get("product_id") or "",
                "public_title": item.get("public_title") or "",
                "sku": sku,
                "legacy_product_id": legacy_id,
                "legacy_raw_code": item.get("legacy_raw_code") or "",
                "legacy_url": item.get("legacy_url") or "",
                "identity_key": identity_key,
                "title_cleanup_status": item.get("title_cleanup_status") or "",
                "review_flags": item.get("review_flags") or "",
            }
        )
    with csv_path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=columns)
        writer.writeheader()
        writer.writerows(normalized)
    os.chmod(csv_path, 0o600)
    lines = [
        "گزارش هویت محصولات Draft رادمان",
        "wp_id | عنوان عمومی | SKU | legacy_product_id | legacy_raw_code | legacy_url | identity_key | وضعیت پاکسازی | بازبینی",
    ]
    for row in normalized:
        lines.append(
            "{wp_id} | {public_title} | {sku} | {legacy_product_id} | "
            "{legacy_raw_code} | {legacy_url} | {identity_key} | "
            "{title_cleanup_status} | {review_flags}".format(**row)
        )
    txt_path.write_text("\n".join(lines) + "\n", encoding="utf-8")
    os.chmod(txt_path, 0o600)
    for row in normalized:
        print(
            f"{row['wp_id']} | {row['public_title']} | {row['sku']} | "
            f"{row['legacy_product_id']} | {row['identity_key']} | "
            f"{row['title_cleanup_status']} | {row['review_flags']}"
        )
    return csv_path, txt_path


def latest_manifest(private_dir: Path) -> Path:
    candidates = sorted(
        (private_dir / "legacy-cache" / "runs").glob(
            "excel-import-*/prepared-products.json"
        ),
        reverse=True,
    )
    if not candidates:
        raise ExcelPipelineError(
            "prepared-products.json پیدا نشد؛ ابتدا --fetch-images یا --full-pilot را اجرا کنید"
        )
    return candidates[0]


def read_manifest(path: Path, private_dir: Path) -> Dict[str, Any]:
    resolved = path.expanduser().resolve()
    private_root = private_dir.expanduser().resolve()
    if private_root not in resolved.parents or not resolved.is_file():
        raise ExcelPipelineError("manifest باید فایل موجود زیر RADMAN_PRIVATE_DIR باشد")
    payload = json.loads(resolved.read_text(encoding="utf-8"))
    records = payload.get("products", [])
    if not isinstance(records, list) or not records or len(records) > HARD_MAX_PRODUCTS:
        raise ExcelPipelineError("manifest باید 1..1000 محصول داشته باشد")
    for record in records:
        if record.get("image_status") == "NOT_FETCHED":
            raise ExcelPipelineError(
                "manifest فقط plan است؛ پیش از import باید --fetch-images اجرا شود"
            )
        for raw_path in record.get("selected_import_paths", []):
            media_path = Path(raw_path).expanduser().resolve()
            if private_root not in media_path.parents:
                raise ExcelPipelineError(
                    f"مسیر تصویر import باید زیر RADMAN_PRIVATE_DIR باشد: {media_path}"
                )
    return payload


def prepare_from_excel(
    excel_file: Path,
    max_products: int,
) -> Tuple[List[Dict[str, Any]], Dict[str, int], List[str]]:
    all_records, warnings = load_excel_records(excel_file)
    selected, selection = select_newest(all_records, max_products)
    return selected, selection, warnings


def print_pricing_preview(
    records: Sequence[Mapping[str, Any]], selection: Mapping[str, int]
) -> None:
    print("=" * 118)
    print("RADMAN EXCEL IMPORT — SELECTION + PRICING PREVIEW (NO WP WRITE)")
    print(
        f"eligible={selection.get('eligible_rows', 0)} selected={len(records)} "
        "sort=legacy_id DESC currency=IRT/Toman"
    )
    print("ID     SKU          SRC                  WEIGHT    EXCEL       COMPUTED    FINAL       PRICE_SOURCE")
    print("-" * 118)
    for record in records[:20]:
        print(
            f"{int(record['legacy_id']):<6} "
            f"{str(record['sku']):<12} "
            f"{str(record['sku_source']):<20} "
            f"{str(record.get('weight_grams') or '-'):>8} "
            f"{str(record.get('excel_price_toman') or '-'):>11} "
            f"{str(record.get('computed_price_toman') or '-'):>11} "
            f"{str(record.get('final_price_toman') or '-'):>11} "
            f"{record.get('price_source')}"
        )
    if len(records) > 20:
        print(f"... {len(records) - 20} additional selected rows are in the CSV/TXT report")
    print("PLAN PREVIEW ONLY — no WordPress product was changed.")
    print("=" * 118)


def inspect_excel(excel_file: Path, max_products: int) -> None:
    records, selection, warnings = prepare_from_excel(excel_file, max_products)
    identifiers = [int(record["legacy_id"]) for record in records]
    print("RADMAN Excel import inspection — READ ONLY")
    print(f"  file: {excel_file}")
    print(f"  sheet: {SHEET_NAME}")
    print(f"  total parsed: {selection['excel_rows']}")
    print(f"  eligible: {selection['eligible_rows']}")
    print(f"  selected cap: {max_products}")
    print(f"  selected: {len(records)}")
    if identifiers:
        print(f"  selected ID range: {min(identifiers)}..{max(identifiers)}")
        print(f"  first/newest IDs: {', '.join(str(value) for value in identifiers[:5])}")
    print("  ordering: ID DESCENDING (higher ID is newer)")
    print(f"  warnings: {len(warnings)}")
    print(f"  API slot: {API_SLOT}")
    print("  mutation: NONE")


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    modes = parser.add_mutually_exclusive_group(required=True)
    modes.add_argument("--inspect", action="store_true")
    modes.add_argument("--plan", action="store_true")
    modes.add_argument("--fetch-images", action="store_true")
    modes.add_argument("--import-drafts", action="store_true")
    modes.add_argument("--enrich-existing", action="store_true")
    modes.add_argument("--identity-report", action="store_true")
    modes.add_argument("--full-pilot", action="store_true")
    parser.add_argument("--excel", type=Path, default=DEFAULT_EXCEL_FILE)
    parser.add_argument("--max-products", type=int, default=DEFAULT_MAX_PRODUCTS)
    parser.add_argument("--private-dir", type=Path, required=True)
    parser.add_argument("--manifest", type=Path)
    parser.add_argument("--wp-path", default=os.environ.get("WP_PATH", ""))
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    private_dir = args.private_dir.expanduser().resolve()
    if not private_dir.is_absolute() or "public_html" in str(private_dir):
        print("[ERROR] RADMAN_PRIVATE_DIR must be private and absolute", file=sys.stderr)
        return 2
    if args.max_products < 1 or args.max_products > HARD_MAX_PRODUCTS:
        print(f"[ERROR] MAX_PRODUCTS must be 1..{HARD_MAX_PRODUCTS}", file=sys.stderr)
        return 2
    try:
        if args.identity_report:
            require_readonly_wp_environment(args.wp_path)
            rows = ExcelDraftGateway(args.wp_path).list_draft_legacy_products(
                args.max_products
            )
            csv_path, txt_path = write_identity_report(rows, private_dir)
            print(f"[IDENTITY REPORT] {csv_path}")
            print(f"[IDENTITY REPORT] {txt_path}")
            print("[SAFETY] Read-only; standard WooCommerce SKU search remains enabled.")
            return 0

        if args.inspect:
            inspect_excel(args.excel, args.max_products)
            return 0

        if args.enrich_existing:
            require_import_environment(args.wp_path)
            all_records, excel_warnings = load_excel_records(args.excel)
            gateway = ExcelDraftGateway(args.wp_path)
            matched, missing = prepare_existing_enrichment(
                all_records, gateway, args.max_products
            )
            if not matched and not missing:
                raise ExcelPipelineError(
                    "هیچ Draft دارای meta legacy_product_id برای enrichment پیدا نشد"
                )
            run_stamp = now_slug()
            run_dir = (
                private_dir
                / "legacy-cache"
                / "runs"
                / f"excel-enrich-{run_stamp}"
            )
            enriched = fetch_specs_for_records(
                matched,
                run_dir=run_dir,
            )
            actions = enrich_existing_records(enriched, gateway)
            actions.extend(missing)
            selection = {
                "excel_rows": len(all_records),
                "eligible_rows": len(matched),
                "selected_rows": len(matched),
                "inactive_skipped": 0,
                "unavailable_skipped": 0,
            }
            payload = {
                "generated_at": datetime.now(TEHRAN).isoformat(),
                "run_stamp": run_stamp,
                "pipeline_version": PIPELINE_VERSION,
                "mode": "enrich-existing",
                "excel_file": str(args.excel.expanduser().resolve()),
                "excel_warnings": excel_warnings,
                "products": enriched,
                "actions": actions,
            }
            write_json_atomic(run_dir / "enrichment-results.json", payload)
            write_json_atomic(run_dir / "import-actions.json", actions)
            write_reports(enriched, run_dir, run_stamp, selection)
            print(
                f"[ENRICH] matched={len(enriched)} missing_excel={len(missing)} "
                "Draft descriptions/meta updated idempotently"
            )
            return 0

        if args.import_drafts:
            require_import_environment(args.wp_path)
            manifest_path = args.manifest or latest_manifest(private_dir)
            payload = read_manifest(manifest_path, private_dir)
            records = payload["products"]
            actions = import_records(records, ExcelDraftGateway(args.wp_path))
            payload["products"] = records
            payload["import_actions"] = actions
            run_dir = manifest_path.resolve().parent
            run_stamp = str(payload.get("run_stamp") or now_slug())
            write_json_atomic(run_dir / "prepared-products.json", payload)
            write_json_atomic(run_dir / "import-actions.json", actions)
            write_reports(
                records,
                run_dir,
                run_stamp,
                payload.get("selection_summary", {}),
            )
            print(f"[IMPORT] actions={len(actions)}; all creations are Draft")
            return 0

        records, selection, warnings = prepare_from_excel(
            args.excel, args.max_products
        )
        if args.plan or args.full_pilot:
            print_pricing_preview(records, selection)
        run_stamp = now_slug()
        run_dir = (
            private_dir
            / "legacy-cache"
            / "runs"
            / f"excel-import-{run_stamp}"
        )
        if args.fetch_images or args.full_pilot:
            records = fetch_images_for_records(
                records,
                private_dir=private_dir,
                run_dir=run_dir,
            )
        payload = {
            "generated_at": datetime.now(TEHRAN).isoformat(),
            "run_stamp": run_stamp,
            "pipeline_version": PIPELINE_VERSION,
            "excel_file": str(args.excel.expanduser().resolve()),
            "sheet": SHEET_NAME,
            "sort": "legacy_id DESC",
            "max_products": args.max_products,
            "selection_summary": selection,
            "excel_warnings": warnings,
            "api_slot": str(API_SLOT),
            "products": records,
        }
        manifest_path = run_dir / "prepared-products.json"
        write_json_atomic(manifest_path, payload)
        csv_path, txt_path = write_reports(
            records, run_dir, run_stamp, selection
        )
        print(f"[REPORT] {csv_path}")
        print(f"[REPORT] {txt_path}")
        print(f"[MANIFEST] {manifest_path}")

        if args.full_pilot:
            require_import_environment(args.wp_path)
            actions = import_records(records, ExcelDraftGateway(args.wp_path))
            payload["products"] = records
            payload["import_actions"] = actions
            write_json_atomic(manifest_path, payload)
            write_json_atomic(run_dir / "import-actions.json", actions)
            write_reports(records, run_dir, run_stamp, selection)
            print(f"[IMPORT] actions={len(actions)}; all creations are Draft")
        return 0
    except (ExcelPipelineError, WPCliError, OSError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
