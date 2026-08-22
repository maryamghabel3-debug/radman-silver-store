#!/usr/bin/env python3
"""Read-only analysis of the owner-exported legacy product Excel workbook.

This tool reads one XLSX sheet, computes category/ID/weight/price/stock summaries,
and prints a Persian Markdown report. It has no WordPress, network, image, or
import capability.
"""

from __future__ import annotations

import argparse
import json
import os
import re
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Dict, Iterable, List, Mapping, Optional, Sequence, Tuple
from zoneinfo import ZoneInfo

SHEET_NAME = "همه محصولات"
EXPECTED_ROWS = 3232
EXPECTED_COLUMNS = 29
EXPECTED_CATEGORY_COUNT = 49
WINDOW_SIZE = 1000
MIN_REQUIRED_COLUMNS = 22

COL_ID = 1
COL_TITLE = 2
COL_CATEGORY = 5
COL_PRICE = 9
COL_AVAILABILITY = 12
COL_ACTIVE = 13
COL_WEIGHT = 22
TITLE_HEADER_CANDIDATES = (
    "عنوان محصول",
    "نام محصول",
    "عنوان کالا",
    "نام کالا",
    "عنوان",
    "نام",
)

TEHRAN = ZoneInfo("Asia/Tehran")
_DIGITS = str.maketrans("۰۱۲۳۴۵۶۷۸۹٠١٢٣٤٥٦٧٨٩", "01234567890123456789")

GEMSTONE_PATTERNS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("عقیق", ("عقیق",)),
    ("فیروزه", ("فیروزه",)),
    ("حدید", ("حدید",)),
    ("در نجف", ("در نجف", "دُر نجف", "درّ نجف")),
    ("یاقوت", ("یاقوت",)),
    ("زمرد", ("زمرد",)),
    ("زبرجد", ("زبرجد",)),
    ("اونیکس", ("اونیکس", "اونکس")),
    ("شرف‌الشمس", ("شرف الشمس", "شرف‌الشمس")),
    ("نگین/سنگ", ("نگین", "سنگ")),
)
AUDIENCE_PATTERNS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("مردانه", ("مردانه",)),
    ("زنانه", ("زنانه",)),
    ("بچگانه", ("بچگانه", "کودک")),
    ("اسپرت/مشترک", ("اسپرت", "مشترک", "یونیسکس")),
)
STYLE_PATTERNS: Tuple[Tuple[str, Tuple[str, ...]], ...] = (
    ("خطی", ("خطی", "خط نوشته", "خط‌نوشته")),
    ("قلم‌زنی", ("قلمزنی", "قلم‌زنی")),
    ("شبکه", ("شبکه", "شبکه‌ای")),
    ("رکاب", ("رکاب",)),
    ("حرز", ("حرز",)),
    ("دست‌ساز", ("دست ساز", "دست‌ساز")),
)


class CatalogAnalysisError(RuntimeError):
    """Raised when the workbook cannot be safely analyzed."""


@dataclass(frozen=True)
class ProductRow:
    excel_row: int
    legacy_id: int
    title: str
    category: Optional[str]
    price: Optional[Decimal]
    weight: Optional[Decimal]
    weight_present: bool
    availability: str
    active: Optional[bool]


@dataclass(frozen=True)
class NumericStats:
    count: int
    missing_or_invalid: int
    minimum: Optional[Decimal]
    maximum: Optional[Decimal]
    average: Optional[Decimal]


@dataclass(frozen=True)
class WindowStats:
    label: str
    count: int
    id_min: Optional[int]
    id_max: Optional[int]
    price: NumericStats
    weight_present: int
    weight_missing: int
    weight_coverage_percent: Decimal


def normalize_text(value: Any) -> str:
    return re.sub(
        r"\s+",
        " ",
        str(value if value is not None else "")
        .translate(_DIGITS)
        .replace("ي", "ی")
        .replace("ك", "ک")
        .replace("\u200c", " "),
    ).strip()


def display_text(value: Any) -> str:
    """Trim layout whitespace while preserving the category's rendered characters."""
    return re.sub(
        r"[ \t\r\n]+", " ", str(value if value is not None else "")
    ).strip()


def is_blank(value: Any) -> bool:
    if value is None:
        return True
    text = normalize_text(value)
    return text in {"", "-", "—", "–", "null", "none", "nan"}


def parse_decimal(value: Any, *, decimal_comma: bool = False) -> Optional[Decimal]:
    if is_blank(value) or isinstance(value, bool):
        return None
    if isinstance(value, (int, float, Decimal)):
        try:
            parsed = Decimal(str(value))
        except InvalidOperation:
            return None
        return parsed if parsed.is_finite() else None

    text = normalize_text(value).lower()
    text = (
        text.replace("تومان", "")
        .replace("ریال", "")
        .replace("گرم", "")
        .replace("٫", ".")
        .replace("٬", "")
        .replace(" ", "")
    )
    if decimal_comma and "." not in text and re.fullmatch(r"[-+]?[0-9]+,[0-9]{1,3}", text):
        text = text.replace(",", ".")
    else:
        text = text.replace(",", "")
    match = re.fullmatch(r"[-+]?[0-9]+(?:\.[0-9]+)?", text)
    if not match:
        number = re.search(r"[-+]?[0-9]+(?:\.[0-9]+)?", text)
        if not number:
            return None
        text = number.group(0)
    try:
        parsed = Decimal(text)
    except InvalidOperation:
        return None
    return parsed if parsed.is_finite() else None


def parse_id(value: Any) -> Optional[int]:
    parsed = parse_decimal(value)
    if parsed is None or parsed <= 0 or parsed != parsed.to_integral_value():
        return None
    return int(parsed)


def parse_active(value: Any) -> Optional[bool]:
    parsed = parse_decimal(value)
    if parsed is not None:
        if parsed == 1:
            return True
        if parsed == 0:
            return False
    text = normalize_text(value).casefold()
    if text in {"true", "yes", "فعال", "بله"}:
        return True
    if text in {"false", "no", "غیرفعال", "خیر"}:
        return False
    return None


def normalize_availability(value: Any) -> str:
    numeric = parse_decimal(value)
    if numeric == 1:
        return "موجود"
    if numeric == 0:
        return "ناموجود"
    text = normalize_text(value).casefold()
    compact = text.replace(" ", "")
    if compact in {"موجود", "instock", "in_stock", "1", "true", "yes"}:
        return "موجود"
    if compact in {
        "ناموجود",
        "ناموجوددرانبار",
        "outofstock",
        "out_of_stock",
        "0",
        "false",
        "no",
    }:
        return "ناموجود"
    return text or "نامشخص"


def _cell(values: Sequence[Any], one_based_column: int) -> Any:
    index = one_based_column - 1
    return values[index] if index < len(values) else None


def resolve_title_column(headers: Sequence[str]) -> int:
    normalized_headers = [normalize_text(header) for header in headers]
    for candidate in TITLE_HEADER_CANDIDATES:
        normalized_candidate = normalize_text(candidate)
        if normalized_candidate in normalized_headers:
            return normalized_headers.index(normalized_candidate) + 1
    for index, header in enumerate(normalized_headers, start=1):
        if "عنوان" in header or ("نام" in header and "محصول" in header):
            return index
    return COL_TITLE


def load_catalog(
    excel_path: Path,
    *,
    sheet_name: str = SHEET_NAME,
    header_row: int = 1,
) -> Tuple[List[ProductRow], List[str], List[str]]:
    """Load valid product rows without modifying the workbook."""
    try:
        from openpyxl import load_workbook
    except ImportError as exc:
        raise CatalogAnalysisError(
            "openpyxl is required: python3 -m pip install --user openpyxl"
        ) from exc

    path = excel_path.expanduser().resolve()
    if not path.is_file():
        raise CatalogAnalysisError(f"فایل Excel پیدا نشد: {path}")
    if path.suffix.lower() != ".xlsx":
        raise CatalogAnalysisError("فقط فرمت .xlsx پشتیبانی می‌شود")
    if header_row < 1:
        raise CatalogAnalysisError("header row must be >= 1")

    try:
        workbook = load_workbook(
            path, read_only=True, data_only=True, keep_links=False
        )
    except Exception as exc:
        raise CatalogAnalysisError(f"خواندن فایل Excel ناموفق بود: {exc}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            available = "، ".join(workbook.sheetnames)
            raise CatalogAnalysisError(
                f"شیت «{sheet_name}» پیدا نشد؛ شیت‌های موجود: {available}"
            )
        worksheet = workbook[sheet_name]
        if worksheet.max_column < MIN_REQUIRED_COLUMNS:
            raise CatalogAnalysisError(
                f"شیت فقط {worksheet.max_column} ستون دارد؛ حداقل {MIN_REQUIRED_COLUMNS} لازم است"
            )

        raw_header = next(
            worksheet.iter_rows(
                min_row=header_row,
                max_row=header_row,
                values_only=True,
            ),
            tuple(),
        )
        headers = [normalize_text(value) for value in raw_header]
        rows: List[ProductRow] = []
        warnings: List[str] = []
        if worksheet.max_column != EXPECTED_COLUMNS:
            warnings.append(
                f"تعداد ستون واقعی {worksheet.max_column} است؛ انتظار مالک {EXPECTED_COLUMNS} ستون بود"
            )
        title_column = resolve_title_column(headers)
        if title_column != COL_TITLE:
            warnings.append(
                f"ستون عنوان از header تشخیص داده شد: ستون {title_column} (fallback تعریف‌شده {COL_TITLE})"
            )
        seen_ids: Counter[int] = Counter()
        for excel_row, values in enumerate(
            worksheet.iter_rows(min_row=header_row + 1, values_only=True),
            start=header_row + 1,
        ):
            if all(is_blank(value) for value in values):
                continue
            legacy_id = parse_id(_cell(values, COL_ID))
            if legacy_id is None:
                warnings.append(
                    f"ردیف {excel_row}: شناسه قدیمی معتبر در ستون {COL_ID} ندارد و رد شد"
                )
                continue
            category_text = display_text(_cell(values, COL_CATEGORY))
            weight_raw = _cell(values, COL_WEIGHT)
            price = parse_decimal(_cell(values, COL_PRICE))
            weight = parse_decimal(weight_raw, decimal_comma=True)
            if price is not None and price < 0:
                warnings.append(f"ردیف {excel_row}: قیمت منفی نامعتبر است")
                price = None
            if weight is not None and weight <= 0:
                warnings.append(f"ردیف {excel_row}: وزن صفر/منفی نامعتبر است")
                weight = None
            rows.append(
                ProductRow(
                    excel_row=excel_row,
                    legacy_id=legacy_id,
                    title=display_text(_cell(values, title_column)) or "(بدون عنوان)",
                    category=category_text or None,
                    price=price,
                    weight=weight,
                    weight_present=not is_blank(weight_raw),
                    availability=normalize_availability(
                        _cell(values, COL_AVAILABILITY)
                    ),
                    active=parse_active(_cell(values, COL_ACTIVE)),
                )
            )
            seen_ids[legacy_id] += 1
    finally:
        workbook.close()

    if not rows:
        raise CatalogAnalysisError("هیچ ردیف محصول معتبر در شیت پیدا نشد")
    duplicates = sorted(identifier for identifier, count in seen_ids.items() if count > 1)
    if duplicates:
        sample = ", ".join(str(value) for value in duplicates[:20])
        warnings.append(
            f"{len(duplicates)} شناسه تکراری دیده شد (نمونه: {sample})"
        )
    if len(headers) >= COL_CATEGORY:
        expected_header = "نام دسته اصلی"
        actual = headers[COL_CATEGORY - 1]
        if actual and expected_header not in actual:
            warnings.append(
                f"عنوان ستون {COL_CATEGORY} «{actual}» است؛ انتظار «{expected_header}» بود"
            )
    return rows, headers, warnings


def numeric_stats(values: Iterable[Optional[Decimal]]) -> NumericStats:
    materialized = list(values)
    valid = [value for value in materialized if value is not None]
    if not valid:
        return NumericStats(0, len(materialized), None, None, None)
    return NumericStats(
        count=len(valid),
        missing_or_invalid=len(materialized) - len(valid),
        minimum=min(valid),
        maximum=max(valid),
        average=sum(valid, Decimal("0")) / Decimal(len(valid)),
    )


def coverage_percent(present: int, total: int) -> Decimal:
    if total <= 0:
        return Decimal("0")
    return (Decimal(present) * Decimal("100")) / Decimal(total)


def window_stats(label: str, rows: Sequence[ProductRow]) -> WindowStats:
    present = sum(row.weight_present for row in rows)
    identifiers = [row.legacy_id for row in rows]
    return WindowStats(
        label=label,
        count=len(rows),
        id_min=min(identifiers) if identifiers else None,
        id_max=max(identifiers) if identifiers else None,
        price=numeric_stats(row.price for row in rows),
        weight_present=present,
        weight_missing=len(rows) - present,
        weight_coverage_percent=coverage_percent(present, len(rows)),
    )


def _contains_any(text: str, variants: Sequence[str]) -> bool:
    normalized = normalize_text(text).casefold()
    return any(normalize_text(value).casefold() in normalized for value in variants)


def matched_labels(
    category: str,
    patterns: Sequence[Tuple[str, Tuple[str, ...]]],
) -> List[str]:
    return [label for label, variants in patterns if _contains_any(category, variants)]


def infer_taxonomy(category: str) -> Tuple[str, str, List[str]]:
    normalized = normalize_text(category)
    audiences = matched_labels(normalized, AUDIENCE_PATTERNS)
    gemstones = matched_labels(normalized, GEMSTONE_PATTERNS)
    styles = matched_labels(normalized, STYLE_PATTERNS)
    audience = audiences[0] if audiences else "عمومی/نیازمند تعیین"

    if "انگشتر" in normalized:
        parent, child = "انگشتر", audience
    elif "دستبند" in normalized:
        parent, child = "دستبند", audience
    elif "گردنبند" in normalized or "سینه ریز" in normalized:
        parent, child = "گردنبند و آویز", "گردنبند"
    elif any(token in normalized for token in ("مدال", "آویز", "اویز")):
        parent, child = "گردنبند و آویز", "مدال و آویز"
    elif "زنجیر" in normalized:
        parent, child = "گردنبند و آویز", "زنجیر"
    elif "گوشواره" in normalized:
        parent, child = "گوشواره", audience
    elif re.search(r"(?:^|\s)(?:نیم\s*ست|ست)(?:\s|$)", normalized):
        parent, child = "ست و نیم‌ست", audience
    elif "نگین" in normalized or "سنگ" in normalized or gemstones:
        parent = "سنگ و نگین"
        child = gemstones[0] if gemstones else "سایر سنگ‌ها"
    else:
        parent, child = "سایر/نیازمند بررسی", "بررسی دستی"

    facets = list(dict.fromkeys([*audiences, *gemstones, *styles]))
    return parent, child, facets


def pattern_summary(
    category_counts: Mapping[str, int],
    patterns: Sequence[Tuple[str, Tuple[str, ...]]],
) -> List[Dict[str, Any]]:
    result = []
    for label, variants in patterns:
        matching = [
            (category, count)
            for category, count in category_counts.items()
            if _contains_any(category, variants)
        ]
        result.append(
            {
                "pattern": label,
                "category_count": len(matching),
                "product_count": sum(count for _, count in matching),
                "categories": [category for category, _ in matching],
            }
        )
    return result


def analyze_catalog(
    rows: Sequence[ProductRow],
    *,
    headers: Sequence[str] = tuple(),
    warnings: Sequence[str] = tuple(),
    expected_rows: int = EXPECTED_ROWS,
    expected_category_count: int = EXPECTED_CATEGORY_COUNT,
) -> Dict[str, Any]:
    sorted_ascending = sorted(rows, key=lambda row: (row.legacy_id, row.excel_row))
    sorted_descending = sorted(rows, key=lambda row: (-row.legacy_id, row.excel_row))
    category_counts = Counter(row.category for row in rows if row.category)
    missing_category = sum(row.category is None for row in rows)
    weight_present = sum(row.weight_present for row in rows)
    valid_weight = sum(row.weight is not None for row in rows)
    availability = Counter(row.availability for row in rows)
    active = Counter(row.active for row in rows)

    mappings = []
    tree_counts: Dict[str, Counter[str]] = defaultdict(Counter)
    for category, count in sorted(
        category_counts.items(), key=lambda item: (-item[1], item[0])
    ):
        parent, child, facets = infer_taxonomy(category)
        tree_counts[parent][child] += count
        mappings.append(
            {
                "legacy_category": category,
                "product_count": count,
                "parent": parent,
                "child": child,
                "facets": facets,
            }
        )

    low_window = sorted_ascending[: min(WINDOW_SIZE, len(sorted_ascending))]
    high_window = sorted_descending[: min(WINDOW_SIZE, len(sorted_descending))]
    analysis = {
        "total_rows": len(rows),
        "expected_rows": expected_rows,
        "column_count": len(headers),
        "expected_columns": EXPECTED_COLUMNS,
        "headers": list(headers),
        "warnings": list(warnings),
        "category_distribution": [
            {"category": category, "count": count}
            for category, count in sorted(
                category_counts.items(), key=lambda item: (-item[1], item[0])
            )
        ],
        "category_count": len(category_counts),
        "expected_category_count": expected_category_count,
        "missing_category": missing_category,
        "id_min": sorted_ascending[0].legacy_id,
        "id_max": sorted_descending[0].legacy_id,
        "lowest_id_products": [asdict(row) for row in sorted_ascending[:5]],
        "highest_id_products": [asdict(row) for row in sorted_descending[:5]],
        "weight": {
            "present": weight_present,
            "missing": len(rows) - weight_present,
            "coverage_percent": coverage_percent(weight_present, len(rows)),
            "valid_numeric": valid_weight,
            "present_but_invalid": weight_present - valid_weight,
        },
        "price": asdict(numeric_stats(row.price for row in rows)),
        "lowest_1000": asdict(window_stats("1000 شناسه پایین‌تر", low_window)),
        "highest_1000": asdict(window_stats("1000 شناسه بالاتر", high_window)),
        "active": {
            "active": active[True],
            "inactive": active[False],
            "unknown": active[None],
        },
        "availability": dict(
            sorted(availability.items(), key=lambda item: (-item[1], item[0]))
        ),
        "patterns": {
            "gemstones": pattern_summary(category_counts, GEMSTONE_PATTERNS),
            "audience": pattern_summary(category_counts, AUDIENCE_PATTERNS),
            "style": pattern_summary(category_counts, STYLE_PATTERNS),
        },
        "taxonomy_tree": {
            parent: dict(
                sorted(children.items(), key=lambda item: (-item[1], item[0]))
            )
            for parent, children in sorted(tree_counts.items())
        },
        "taxonomy_mappings": mappings,
    }
    return analysis


def format_decimal(value: Optional[Decimal], places: int = 0) -> str:
    if value is None:
        return "—"
    quantum = Decimal("1") if places == 0 else Decimal("1").scaleb(-places)
    rounded = value.quantize(quantum, rounding=ROUND_HALF_UP)
    if places == 0:
        return f"{int(rounded):,}"
    return f"{rounded:,.{places}f}"


def _stats_lines(stats: Mapping[str, Any], *, prefix: str = "") -> List[str]:
    return [
        f"{prefix}- تعداد قیمت معتبر: {stats['count']:,}",
        f"{prefix}- قیمت خالی/نامعتبر: {stats['missing_or_invalid']:,}",
        f"{prefix}- کمینه: {format_decimal(stats['minimum'])} تومان",
        f"{prefix}- بیشینه: {format_decimal(stats['maximum'])} تومان",
        f"{prefix}- میانگین: {format_decimal(stats['average'])} تومان",
    ]


def escape_markdown_cell(value: Any) -> str:
    return str(value).replace("|", "\\|").replace("\n", " ")


def _product_table(products: Sequence[Mapping[str, Any]]) -> List[str]:
    lines = [
        "| ID | عنوان | قیمت (تومان) | وزن (گرم) |",
        "|---:|---|---:|---:|",
    ]
    for product in products:
        lines.append(
            "| {id} | {title} | {price} | {weight} |".format(
                id=product["legacy_id"],
                title=escape_markdown_cell(product["title"]),
                price=format_decimal(product["price"]),
                weight=format_decimal(product["weight"], 3),
            )
        )
    return lines


def render_report(
    analysis: Mapping[str, Any],
    *,
    source_path: Path,
    sheet_name: str,
) -> str:
    now = datetime.now(TEHRAN).isoformat(timespec="seconds")
    category_delta = analysis["category_count"] - analysis["expected_category_count"]
    row_delta = analysis["total_rows"] - analysis["expected_rows"]
    weight = analysis["weight"]
    lines: List[str] = [
        "# تحلیل کاتالوگ Excel محصولات قدیمی رادمان",
        "",
        f"- زمان تولید: `{now}` (`Asia/Tehran`)",
        f"- فایل منبع (فقط خواندنی): `{source_path}`",
        f"- شیت: `{sheet_name}`",
        "- دامنه مأموریت: فقط تحلیل؛ بدون WordPress، تصویر، import یا network",
        "",
        "## نتیجه مدیریتی و فرض ترتیب ID",
        "",
        f"- ردیف محصول معتبر: **{analysis['total_rows']:,}** (انتظار مالک: {analysis['expected_rows']:,}؛ اختلاف {row_delta:+,})",
        f"- تعداد ستون: **{analysis['column_count']:,}** (انتظار مالک: {analysis['expected_columns']:,})",
        f"- بازه ID: **{analysis['id_min']:,} تا {analysis['id_max']:,}**",
        "- **فرض اعلام‌شده مالک:** ID پایین‌تر یعنی محصول جدیدتر؛ ID بالاتر یعنی محصول قدیمی‌تر.",
        "- این گزارش فقط دو سر بازه را نشان می‌دهد و از خود ID تاریخ را استنباط نمی‌کند؛ مالک باید این فرض را با محصولات نمونه زیر تأیید کند.",
        "",
        "## ۱) توزیع دسته‌ها",
        "",
        f"- تعداد دسته یکتای غیرخالی: **{analysis['category_count']:,}** (انتظار تقریبی: {analysis['expected_category_count']:,}؛ اختلاف {category_delta:+,})",
        f"- محصول بدون دسته: **{analysis['missing_category']:,}**",
        "",
        "| رتبه | نام دقیق دسته قدیمی | تعداد محصول |",
        "|---:|---|---:|",
    ]
    for rank, item in enumerate(analysis["category_distribution"], start=1):
        lines.append(
            f"| {rank} | {escape_markdown_cell(item['category'])} | {item['count']:,} |"
        )

    lines.extend(
        [
            "",
            "## ۲) بازه ID و نمونه‌های دو سر کاتالوگ",
            "",
            "### ۵ ID پایین‌تر — طبق فرض مالک، جدیدترین‌ها",
            "",
            *_product_table(analysis["lowest_id_products"]),
            "",
            "### ۵ ID بالاتر — طبق فرض مالک، قدیمی‌ترین‌ها",
            "",
            *_product_table(analysis["highest_id_products"]),
            "",
            "## ۳) پوشش وزن (ستون ۲۲)",
            "",
            f"- دارای سلول وزن غیرخالی: **{weight['present']:,}**",
            f"- وزن خالی: **{weight['missing']:,}**",
            f"- پوشش: **{format_decimal(weight['coverage_percent'], 2)}٪**",
            f"- وزن عددی معتبر: **{weight['valid_numeric']:,}**",
            f"- سلول وزن غیرخالی ولی غیرقابل‌تبدیل به عدد: **{weight['present_but_invalid']:,}**",
            "",
            "## ۴) آمار قیمت",
            "",
            *_stats_lines(analysis["price"]),
        ]
    )

    for key, heading in (
        ("lowest_1000", "۱۰۰۰ محصول با ID پایین‌تر — طبق فرض مالک، جدیدتر"),
        ("highest_1000", "۱۰۰۰ محصول با ID بالاتر — طبق فرض مالک، قدیمی‌تر"),
    ):
        window = analysis[key]
        lines.extend(
            [
                "",
                f"### {heading}",
                "",
                f"- تعداد: {window['count']:,}",
                f"- بازه ID در نمونه: {window['id_min']} تا {window['id_max']}",
                *_stats_lines(window["price"]),
                f"- وزن موجود: {window['weight_present']:,}",
                f"- وزن خالی: {window['weight_missing']:,}",
                f"- پوشش وزن: {format_decimal(window['weight_coverage_percent'], 2)}٪",
            ]
        )

    availability = analysis["availability"]
    active = analysis["active"]
    lines.extend(
        [
            "",
            "## ۵) فعال بودن و وضعیت موجودی",
            "",
            f"- فعال (`COL 13 = 1`): **{active['active']:,}**",
            f"- غیرفعال (`COL 13 = 0`): **{active['inactive']:,}**",
            f"- وضعیت فعال نامشخص: **{active['unknown']:,}**",
            f"- موجود (`COL 12`): **{availability.get('موجود', 0):,}**",
            f"- ناموجود (`COL 12`): **{availability.get('ناموجود', 0):,}**",
            "",
            "### سایر مقادیر دقیق ستون موجودی",
            "",
            "| مقدار نرمال‌شده | تعداد |",
            "|---|---:|",
        ]
    )
    for value, count in availability.items():
        lines.append(f"| {escape_markdown_cell(value)} | {count:,} |")

    lines.extend(
        [
            "",
            "## ۶) الگوهای نام دسته",
            "",
            "این جدول تعداد دسته‌های یکتا و تعداد محصولات درگیر هر الگو را نشان می‌دهد. یک دسته ممکن است در چند الگو شمرده شود.",
            "",
            "| گروه | الگو | دسته‌های منطبق | محصولات منطبق |",
            "|---|---|---:|---:|",
        ]
    )
    for group_key, group_label in (
        ("gemstones", "سنگ/نگین"),
        ("audience", "مخاطب"),
        ("style", "سبک"),
    ):
        for item in analysis["patterns"][group_key]:
            lines.append(
                f"| {group_label} | {item['pattern']} | {item['category_count']:,} | {item['product_count']:,} |"
            )

    lines.extend(
        [
            "",
            "## ۷) پیشنهاد taxonomy تمیز دو سطحی",
            "",
            "**قاعده پیشنهادی:** سطح اول فقط نوع محصول باشد؛ سطح دوم زیرنوع/مخاطب باشد. نوع سنگ، سبک «خطی»، رکاب، قلم‌زنی و ویژگی‌های مشابه به attribute/filter یا tag منتقل شوند تا دسته‌های ترکیبی تکراری ساخته نشوند.",
            "",
            "### درخت پیشنهادی بر اساس نام‌های موجود",
            "",
        ]
    )
    for parent, children in analysis["taxonomy_tree"].items():
        lines.append(f"- **{parent}**")
        for child, count in children.items():
            lines.append(f"  - {child}: {count:,} محصول")

    lines.extend(
        [
            "",
            "### نگاشت پیشنهادی همه دسته‌های قدیمی",
            "",
            "| دسته قدیمی | تعداد | سطح ۱ | سطح ۲ | attribute/tag پیشنهادی |",
            "|---|---:|---|---|---|",
        ]
    )
    for item in analysis["taxonomy_mappings"]:
        facets = "، ".join(item["facets"]) or "—"
        lines.append(
            "| {legacy} | {count:,} | {parent} | {child} | {facets} |".format(
                legacy=escape_markdown_cell(item["legacy_category"]),
                count=item["product_count"],
                parent=item["parent"],
                child=item["child"],
                facets=facets,
            )
        )

    lines.extend(
        [
            "",
            "## هشدارها و کنترل کیفیت",
            "",
        ]
    )
    if analysis["warnings"]:
        lines.extend(f"- {warning}" for warning in analysis["warnings"])
    else:
        lines.append("- هشدار ساختاری ثبت نشد.")
    lines.extend(
        [
            "",
            "## تصمیم لازم از مالک",
            "",
            "1. پنج محصول ID پایین و پنج محصول ID بالا را با سایت قدیمی مقایسه و فرض «ID پایین‌تر = جدیدتر» را تأیید یا رد کند.",
            "2. تعداد دسته‌های یکتا را با انتظار تقریبی ۴۹ مقایسه کند.",
            "3. taxonomy پیشنهادی صرفاً برنامه‌ریزی است؛ هیچ دسته یا محصولی تغییر نکرده است.",
            "",
        ]
    )
    return "\n".join(lines)


def json_safe(value: Any) -> Any:
    if isinstance(value, Decimal):
        return format(value, "f")
    if isinstance(value, Path):
        return str(value)
    if isinstance(value, dict):
        return {key: json_safe(item) for key, item in value.items()}
    if isinstance(value, (list, tuple)):
        return [json_safe(item) for item in value]
    return value


def validate_output_path(path: Path, source_path: Path) -> Path:
    resolved = path.expanduser().resolve()
    if "public_html" in str(resolved):
        raise CatalogAnalysisError("نوشتن گزارش زیر public_html ممنوع است")
    if resolved == source_path.expanduser().resolve():
        raise CatalogAnalysisError("مسیر گزارش نمی‌تواند همان فایل Excel منبع باشد")
    return resolved


def write_atomic(path: Path, content: str) -> None:
    path.parent.mkdir(parents=True, exist_ok=True, mode=0o700)
    temporary = path.with_suffix(path.suffix + ".tmp")
    temporary.write_text(content, encoding="utf-8")
    os.chmod(temporary, 0o600)
    os.replace(temporary, path)


def parse_args(argv: Optional[Sequence[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--excel", type=Path, required=True)
    parser.add_argument("--sheet", default=SHEET_NAME)
    parser.add_argument("--header-row", type=int, default=1)
    parser.add_argument("--expected-rows", type=int, default=EXPECTED_ROWS)
    parser.add_argument(
        "--expected-categories", type=int, default=EXPECTED_CATEGORY_COUNT
    )
    parser.add_argument("--output", type=Path, help="Optional UTF-8 text/Markdown copy")
    parser.add_argument("--json-output", type=Path, help="Optional machine-readable analysis")
    return parser.parse_args(argv)


def main(argv: Optional[Sequence[str]] = None) -> int:
    args = parse_args(argv)
    try:
        rows, headers, warnings = load_catalog(
            args.excel,
            sheet_name=args.sheet,
            header_row=args.header_row,
        )
        analysis = analyze_catalog(
            rows,
            headers=headers,
            warnings=warnings,
            expected_rows=args.expected_rows,
            expected_category_count=args.expected_categories,
        )
        report = render_report(
            analysis,
            source_path=args.excel.expanduser().resolve(),
            sheet_name=args.sheet,
        )
        output_path = (
            validate_output_path(args.output, args.excel) if args.output else None
        )
        json_output_path = (
            validate_output_path(args.json_output, args.excel)
            if args.json_output
            else None
        )
        if output_path and json_output_path and output_path == json_output_path:
            raise CatalogAnalysisError("مسیر گزارش متنی و JSON باید متفاوت باشد")
        if output_path:
            write_atomic(output_path, report + "\n")
        if json_output_path:
            write_atomic(
                json_output_path,
                json.dumps(json_safe(analysis), ensure_ascii=False, indent=2) + "\n",
            )
        print(report)
        return 0
    except (CatalogAnalysisError, OSError, ValueError) as exc:
        print(f"[ERROR] {exc}", file=sys.stderr)
        return 2


if __name__ == "__main__":
    raise SystemExit(main())
